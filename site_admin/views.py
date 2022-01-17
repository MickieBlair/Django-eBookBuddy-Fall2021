from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from django.contrib.auth import authenticate
from django.utils import timezone
import json
import re
from django.conf import settings
from json import dumps
from django.http import JsonResponse, HttpResponse
import datetime
import calendar
import string
import csv
import xlwt
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Fill
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl.utils import get_column_letter
from copy import copy

from users.models import CustomUser, Role, Secondary_Role
from site_admin.models import Mega_Team, Team, User_Log
from site_admin.models import Gender, Language, System_Message, School, Grade
from site_admin.models import Student_Profile, Volunteer_Profile, Staff_Profile
from site_admin.models import Day, Semester, Session_Day_Time, Daily_Session
from site_admin.models import Room_Type, Room, Day_With_Daily_Session
from site_admin.models import Note_Group, Note_Category, Note
from site_admin.models import Reading_Level, Session_Reading_Level, Student_Progress
from site_admin.models import Student_Assessment
from site_admin.models import Upload_CSV, Server_Time
from site_admin.models import Sign_In_By_User
from site_admin.models import Student_Report, Volunteer_Report, Sign_In_By_Day, Attendance_Status
from site_admin.models import Team_Meeting, Day_With_Team_Meeting
from reading_sessions.models import User_Session_Status, End_Session_Evaluation
from reading_sessions.models import Match_Type, Temporary_Match_Type, Temporary_Match
from reading_sessions.models import Scheduled_Match, Match_Status, Match_Status_Option
from reading_sessions.models import Match_Attendance_Record
from reading_sessions.models import Arrival_Description, Follow_Up_Type
from reading_sessions.models import Evaluation_Level, Relational_Engagement
from reading_sessions.models import Incomplete_Evaluation, A_Problem_User
from users.forms import Create_User_Form, User_Update_Form
from site_admin.forms import Create_Semester_Form, Edit_Semester_Form
from site_admin.forms import Create_Session_Day_Time_Form
from site_admin.forms import Edit_Student_Profile_Form, Edit_Volunteer_Profile_Form, Edit_Staff_Profile_Form
from site_admin.forms import Create_Room_Form, Breakout_Rooms_Form
from site_admin.forms import Create_Scheduled_Match_Form, Edit_Scheduled_Match_Form
from site_admin.forms import CSVBulkUploadForm

from PIL import Image, ImageDraw, ImageFont
import os
from io import BytesIO

from django.core.files import File  # you need this somewhere
from django.core.files.storage import default_storage

# Create your views here.
# def base_view(request, *args, **kwargs):
# 	context = {}
# 	context['page_title'] = "Admin Home"
# 	user = request.user
# 	if not user.is_authenticated:
# 		return redirect('must_authenticate')
# 	else:		
# 		if user.role.name == "Staff":
#			context = button_bar_context_additions(context)	
# 			return render(request, "site_admin/admin_home.html", context)
# 		else:
# 			return redirect('access_denied')

def download_time_log_report_by_role(request, role):
	print("Downloading role Logs Report")
	semester = Semester.objects.get(active_semester=True)
	date=timezone.localtime(timezone.now())
	
	
	if role == "All":
		for_users = CustomUser.objects.all().order_by('username')
		title = "All Users -" + semester.name
		a1_cell = "All Users"
		file_str = 'attachment; filename=' + 'All_Users_Time_Report_' +str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	
	elif role == "Students":
		for_users = CustomUser.objects.filter(role__name="Student").order_by('username')
		title = "Student Users -" + semester.name
		a1_cell = "Student Users"
		file_str = 'attachment; filename=' + 'Student_Users_Time_Report_' +str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'

	elif role == "Volunteers":
		for_users = CustomUser.objects.filter(role__name="Volunteer").order_by('username')
		title = "Volunteer Users -" + semester.name
		a1_cell = "Volunteer Users"
		file_str = 'attachment; filename=' + 'Volunteer_Users_Time_Report' +str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'

	elif role == "Staff":
		for_users = CustomUser.objects.filter(role__name="Staff").order_by('username')
		title = "Staff Users -" + semester.name
		a1_cell = "Staff Users"
		file_str = 'attachment; filename=' + 'Staff_Users_Time_Report' +str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'


			
	wb = Workbook()

	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	

	ws.title = title

	ws['A1'] = a1_cell

	ws['I1'] = "User Count"
	ws['I2'] = str(for_users.count())

	for cell in  ws[1]:
		cell.font = Font(bold=True)

	for cell in  ws[2]:
		cell.border = Border(bottom=double)

	ws.append([])

	ws.append(['ID', 'Username', 'Full Name', 'Role', 'Dropped', 'Log Count',
				'Problem User', 'Total Hours', 'Comment'])
	for cell in  ws[4]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=thin)

	
	for member in for_users:
		if member.user_dropped:
			dropped_string = "Yes"
		else:
			dropped_string = ""

		if member.user_sign_ins.problem_user:
			problem_string = "Yes"
		else:
			problem_string = ""

		if member.user_sign_ins.problem_user:

			comment_str = member.connection_problem.all().first().comment
		else:
			comment_str = ""

		ws.append([str(member.id), str(member.username), str(member.full_name),
				str(member.role.name), dropped_string, str(member.user_sign_ins.good_day_logs.count()),
				problem_string, str(member.user_sign_ins.total_hours_good), str(comment_str)])


	dims = {}
	for row in ws.rows:
		for cell in row:
			print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5


	response = HttpResponse(content_type='application/ms-excel')

	# Set the content disposition as attachment and provide the filename
	# response['Content-Disposition'] = 'attachment; filename=evaluations_report.xlsx'
	response['Content-Disposition'] = file_str

	wb.save(response)
	return response

def download_time_log_individual_report(request, member_id):
	print("Downloading Individual Logs Report")
	member = CustomUser.objects.get(id=member_id)
	logs = User_Log.objects.filter(user=member, count_in_total=True).order_by('date')
	semester = Semester.objects.get(active_semester=True)
	final = Sign_In_By_User.objects.get(user=member)

	# style = xlwt.easyxf('pattern: pattern solid, fore_colour light_blue;'
 #                              'font: colour white, bold True;')

	if final.problem_user:
		problem_string = "Yes"
	else:
		problem_string = ""

	if member.user_dropped:
		dropped_string = "Yes"
	else:
		dropped_string = ""	

			
	wb = Workbook()
	
	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	title = member.full_name + "-" + semester.name

	ws.title = title

	ws['A1'] = member.full_name

	ws['C1'] = "Total Hours"
	ws['C2'] = str(final.total_hours_good)

	ws['D1'] = "Total Logs"
	ws['D2'] = str(logs.count())	

	ws['E1'] = "Problem User"
	ws['E2'] = problem_string

	ws['F1'] = "User Dropped"
	ws['F2'] = dropped_string

	for cell in  ws[1]:
		cell.font = Font(bold=True)

	for cell in  ws[2]:
		cell.border = Border(bottom=double)

	ws.append([])

	starting_row = 5

	if member.user_sign_ins.problem_user:
		comment_str = member.connection_problem.all().first().comment		
		
		ws.append([comment_str])
		ws.merge_cells(start_row=starting_row-1, start_column=1, end_row=starting_row-1, end_column=6)
		for cell in  ws[starting_row-1]:
			cell.fill = PatternFill("solid", fgColor="F5B4B3")
		starting_row = starting_row + 1

		ws.append([])
		starting_row = starting_row + 1
		
	
	

	ws.append(['Date', 'Day', 'Room', 'Time In', 'Time Out', 'Duration Minutes'])
	for cell in  ws[starting_row-1]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=thin)

	
	for log in logs:
		date_str = log.date.strftime("%B %d, %Y")

		if log.room:
			room_str = log.room.name
		else:
			room_str = ""

		if log.manually_added:
			manually_added = "*"
		else:
			manually_added = ""


		ws.append([date_str, log.day_of_week, room_str, log.local_time_in_only_str(),
			log.local_time_out_only_str(), log.minute_str()])

		if log.manually_added:
			for cell in  ws[starting_row]:
					cell.fill = PatternFill("solid", fgColor="F5B4B3")

		starting_row = starting_row + 1

	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=' + member.username +'_' +str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	

	dims = {}
	for row in ws.rows:
		for cell in row:
			print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
				# if cell.value == "*":
				# 	cell.fill = PatternFill("solid", fgColor="F5B4B3" )

			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5


	response = HttpResponse(content_type='application/ms-excel')

	# Set the content disposition as attachment and provide the filename
	# response['Content-Disposition'] = 'attachment; filename=evaluations_report.xlsx'
	response['Content-Disposition'] = file_str

	wb.save(response)
	return response

def all_logs_fill_in_time(log, user):
	print(user)
	user_logs = User_Log.objects.filter(user = user)
	log_list = list(user_logs)

	if log.time_in and not log.time_out:
		print("\n\nMissing Time Out", user)		

		for index, item in enumerate(log_list):
			if not item.time_out:
				missing_out_index = index
				next_index_in_list = missing_out_index + 1
				missing_log = log_list[missing_out_index]
				next_log = log_list[next_index_in_list]
				print("\n\n\nMissing", missing_log.time_in.date(), timezone.localtime(missing_log.time_in).date())
				print("Next In List", next_log.time_in.date(), timezone.localtime(next_log.time_in).date())


				if timezone.localtime(missing_log.time_in).date() == timezone.localtime(next_log.time_in).date():
					print("Can Fill In")
					try:
						use_this = None
						for index, item in enumerate(log_list):
							if index == missing_out_index:
								to_fix = item
								print("This one is missing", index, to_fix.id)
							elif index == next_index_in_list:
								print("Use this time IN", index)
								use_this = log_list[next_index_in_list]
								print("use_this", use_this.id, use_this.time_in.date(), use_this.time_in.time())
								to_fix.time_out = use_this.time_in
								to_fix.save()
								time_difference = to_fix.time_out - to_fix.time_in
								duration_seconds = time_difference.seconds
								if duration_seconds == 0:
									print("zero seconds", log.id, log.user, duration_seconds)
									duration_seconds = 1
								to_fix.duration_seconds = duration_seconds
								to_fix.save()

					except Exception as e:
						print("*************No NEXT", e)
				else:
					print("Next not same day")



def final_fill_in_time(log, log_list, user):
	if log.time_in and not log.time_out:
		print("\n\nMissing Time Out", user)
		

		for index, item in enumerate(log_list):
			if not item.time_out:
				missing_out_index = index
				next_index_in_list = missing_out_index + 1
				missing_log = log_list[missing_out_index]
				next_log = log_list[next_index_in_list]
				print("\n\n\nMissing", missing_log.time_in.date())
				print("Next In List", next_log.time_in.date())

				if missing_log.time_in.date() == next_log.time_in.date():
					print("Can Fill In")
					try:
						use_this = None
						for index, item in enumerate(log_list):
							if index == missing_out_index:
								to_fix = item
								# print("This one is missing", index, to_fix.id)
							elif index == next_index_in_list:
								# print("Use this time IN", index)
								use_this = log_list[next_index_in_list]
								# print("use_this", use_this.id, use_this.time_in.date(), use_this.time_in.time())
								to_fix.time_out = use_this.time_in
								to_fix.save()
								time_difference = to_fix.time_out - to_fix.time_in
								duration_seconds = time_difference.seconds
								if duration_seconds == 0:
									print("zero seconds", log.id, log.user, duration_seconds)

								to_fix.save()

					except Exception as e:
						print("*************No NEXT", e)
				else:
					print("Next not same day")

def all_logs_view(request):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.is_superuser:
			context = button_bar_context_additions(context)
			print("\n\n\nProcess All Logs Evaluations")
			# logs = User_Log.objects.all().order_by('day_of_week', '-needs_attention', 'user')
			# logs = User_Log.objects.all().order_by('room_good', 'user')
			display_logs = User_Log.objects.filter(duration_seconds=0)
			

			logs = User_Log.objects.all().order_by('id')
			context['logs'] = logs

			

			

			semester = Semester.objects.get(active_semester=True)
			context['semester'] = semester
			server_time = Server_Time.objects.get(active=True)

			good_days = ["Monday", 'Tuesday', 'Wednesday', 'Thursday']

			# part one

			add_day = False
			if add_day:
				for log in logs:
					day = log.date.weekday()
					print(day)
					day_of_week = calendar.day_name[day]
					print(day_of_week)
					log.day_of_week = str(day_of_week)
					log.save()

			fill_in_time = False
			if fill_in_time:
				for log in logs:
					day = log.date.weekday()
					print(day)
					day_of_week = calendar.day_name[day]
					print(day_of_week)
					log.day_of_week = str(day_of_week)

					if log.duration_seconds == None:
						if log.time_in and log.time_out:
							time_difference = log.time_out - log.time_in							
							duration_seconds = time_difference.seconds
							if duration_seconds == 0:
								print("zero seconds", log.id, log.user, duration_seconds)
								duration_seconds = 1
							log.duration_seconds = duration_seconds	
							log.needs_attention = False					
							log.save()
							# print("Good")

						else:
							log.needs_attention = False
							all_logs_fill_in_time(log, log.user)
							
							log.save()
							# print("Bad")
					else:
						print("ID, duration", log.id, log.duration_seconds)
						if log.duration_seconds == 0:
							log.duration_seconds = 1
							log.save()


			all_missing_time_count = User_Log.objects.filter(time_out = None).order_by('user')
			context['all_missing_time_count'] = all_missing_time_count.count()
			process_all_missing_time = False
			if process_all_missing_time:
				for item in all_missing_time_count:
					print("Missing Out", item.user, item.id)
					all_logs_fill_in_time(item, item.user)


			still_missing_time = User_Log.objects.filter(time_out = None).order_by('user')	
			print("*******STILL MISSING COUNT 1", still_missing_time.count())
			process_still_missing = False
			if process_still_missing:
				for item in still_missing_time:
					print("Missing Out", item.user, item.id)
					all_logs_fill_in_time(item, item.user)

			# Above part one done

			process_orientation_logs = User_Log.objects.filter(room__name="Orientation")

			process_orientation = False
			if process_orientation:
				for log in process_orientation_logs:
					log.processed_completely = True
					log.count_in_total = True
					log.day_good = True
					log.save()


			buddy_admin_good = User_Log.objects.filter(processed_completely=False, user__username="Buddy_Admin")
			process_buddy_admin = False	
			if process_buddy_admin:
				for log in buddy_admin_good:
					log.processed_completely = True
					log.count_in_total = True
					log.save()

			mickie_good = User_Log.objects.filter(processed_completely=False, user__username="Mickie")
			process_mickie = False	
			if process_mickie:
				for log in mickie_good:
					log.processed_completely = True
					log.count_in_total = True
					log.save()




			student_role = Role.objects.get(name="Student")
			student_days =	['Tuesday', 'Wednesday', 'Thursday']		
			process_student_wrong_day_logs = User_Log.objects.filter(processed_completely=False,
																		user__role=student_role)

			process_student_wrong_day = False
			if process_student_wrong_day:
				for log in process_student_wrong_day_logs:
					if log.day_of_week not in student_days:
						print("Student Wrong Day", log.user, log.day_of_week)
						log.day_good = False
						log.processed_completely = True
						log.count_in_total = False
						log.save()
					else:
						print("Student Right Day", log.user, log.day_of_week)
						log.day_good = True
						log.save()

			unprocessed_server_time_student = User_Log.objects.filter(processed_completely=False,
																		user__role=student_role)
			process_server_time_student = False
			server_time_in = server_time.student_start
			server_time_out = server_time.student_end
			

			if process_server_time_student:
				for log in unprocessed_server_time_student:
					log_in = timezone.localtime(log.time_in).time()
					log_out = timezone.localtime(log.time_out).time()
					print("allow in", server_time_in)
					print("allow out", server_time_out)
					print("In", log_in)
					print("Out", log_out)
					in_good = False
					out_good = False
					all_good = False

					if log_in >= server_time_in and log_in <= server_time_out:
						in_good = True
						print("Time in is Good")
						log.time_in_jitsi_open = True
						log.save()

					else:
						in_good = False
						print("Time in is Bad")
						log.time_in_jitsi_open = False
						log.save()

					if log_out >= server_time_in and log_out <= server_time_out:
						out_good = True
						print("Time out is Good")
						log.time_out_jitsi_open = True
						log.save()

					else:
						out_good = False
						print("Time out is Bad")
						log.time_out_jitsi_open = False
						log.save()

					if in_good and out_good:
						all_good = True

					if all_good:
						log.during_jitsi_open = True
						log.processed_completely = True
						log.count_in_total = True
						log.save()
					else:
						log.processed_completely = True
						log.count_in_total = False
						log.save()


			vol_role = Role.objects.get(name="Volunteer")
			vol_days =	['Monday','Tuesday', 'Wednesday', 'Thursday']		
			process_vol_wrong_day_logs = User_Log.objects.filter(processed_completely=False,
																		user__role=vol_role)


			process_vol_wrong_day = False
			if process_vol_wrong_day:
				for log in process_vol_wrong_day_logs:
					if log.day_of_week not in vol_days:
						print("Vol Wrong Day", log.user, log.day_of_week)
						log.day_good = False
						log.processed_completely = True
						log.count_in_total = False
						log.save()
					else:
						print("Vol Right Day", log.user, log.day_of_week)
						log.day_good = True
						log.save()

			monday_vol_logs = User_Log.objects.filter(processed_completely=False,
													user__role=vol_role,
													day_of_week="Monday")

			room_type = Room_Type.objects.get(letter="B")

			process_meeting_only_monday = False
			if process_meeting_only_monday:
				for log in monday_vol_logs:
					if log.room.room_type == room_type:
						print("Room Type Breakout")
						log.processed_completely = True
						log.count_in_total = False
						log.save()
					else:
						print("Other Type")


			unprocessed_server_time_vol = User_Log.objects.filter(processed_completely=False,
																		user__role=vol_role)
			process_server_time_vol = False
			vserver_time_in = server_time.vol_start
			vserver_time_out = server_time.vol_end

			if process_server_time_vol:
				for log in unprocessed_server_time_vol:
					log_in = timezone.localtime(log.time_in).time()
					log_out = timezone.localtime(log.time_out).time()
					print("allow in", vserver_time_in)
					print("allow out", vserver_time_out)
					print("In", log_in)
					print("Out", log_out)
					in_good = False
					out_good = False
					all_good = False

					if log_in >= vserver_time_in and log_in <= vserver_time_out:
						in_good = True
						print("Time in is Good")
						log.time_in_jitsi_open = True
						log.save()

					else:
						in_good = False
						print("Time in is Bad")
						log.time_in_jitsi_open = False
						log.save()

					if log_out >= vserver_time_in and log_out <= vserver_time_out:
						out_good = True
						print("Time out is Good")
						log.time_out_jitsi_open = True
						log.save()

					else:
						out_good = False
						print("Time out is Bad")
						log.time_out_jitsi_open = False
						log.save()

					if in_good and out_good:
						all_good = True

					if all_good:
						log.during_jitsi_open = True
						log.processed_completely = True
						log.count_in_total = True
						log.save()
					else:
						if in_good:
							log.processed_completely = True
							log.count_in_total = True
							log.save()

						elif out_good:
							log.processed_completely = True
							log.count_in_total = True
							log.save()

						else:
							log.processed_completely = True
							log.count_in_total = False
							log.save()


			chrissy_log = User_Log.objects.filter(processed_completely=False, user__username="Chrissy")
			for log in chrissy_log:
				log.processed_completely = True
				log.count_in_total = True
				log.save()

			venuri_log = User_Log.objects.filter(processed_completely=False, user__username="Venuri")
			for log in venuri_log:
				log.processed_completely = True
				log.count_in_total = True
				log.save()

				

			yaejin_log = User_Log.objects.filter(processed_completely=False, user__username="Yaejin")
			for log in yaejin_log:
				log.processed_completely = True
				log.count_in_total = True
				log.save()


			remaining_logs = User_Log.objects.filter(processed_completely=False,
													).order_by('user')

			process_remaining_day = False
			if process_remaining_day:
				for log in remaining_logs:
					if log.day_of_week not in vol_days:
						print("rest Wrong Day", log.user, log.day_of_week)
						log.day_good = False
						log.processed_completely = True
						log.count_in_total = False
						log.save()
					else:
						print("rest Right Day", log.user, log.day_of_week)
						log.day_good = True
						log.save()

			time_remaining_logs = User_Log.objects.filter(processed_completely=False,
													).order_by('user')

			process_remaining_time = False
			rserver_time_in = server_time.vol_start
			rserver_time_out = server_time.vol_end

			if process_remaining_time:
				for log in time_remaining_logs:
					log_in = timezone.localtime(log.time_in).time()
					log_out = timezone.localtime(log.time_out).time()
					print("allow in", rserver_time_in)
					print("allow out", rserver_time_out)
					print("In", log_in)
					print("Out", log_out)
					in_good = False
					out_good = False
					all_good = False

					if log_in >= rserver_time_in and log_in <= rserver_time_out:
						in_good = True
						print("Time in is Good")
						log.time_in_jitsi_open = True
						log.save()

					else:
						in_good = False
						print("Time in is Bad")
						log.time_in_jitsi_open = False
						log.save()

					if log_out >= rserver_time_in and log_out <= rserver_time_out:
						out_good = True
						print("Time out is Good")
						log.time_out_jitsi_open = True
						log.save()

					else:
						out_good = False
						print("Time out is Bad")
						log.time_out_jitsi_open = False
						log.save()

					if in_good and out_good:
						all_good = True

					if all_good:
						log.during_jitsi_open = True
						log.processed_completely = True
						log.count_in_total = True
						log.save()
					else:
						if in_good:
							log.processed_completely = True
							log.count_in_total = True
							log.save()

						if log.room.room_type.letter == "M":
							log.processed_completely = True
							log.count_in_total = False
							log.save()

						if log.room.room_type.letter == "B":
							log.processed_completely = True
							log.count_in_total = False
							log.save()

						full_access = Secondary_Role.objects.get(name="Full Access")

						if full_access in log.user.secondary_roles.all():
							print("\n\n\nFull access", log.user)
							log.processed_completely = True
							log.count_in_total = True
							log.save()
						else:
							# print("\n\n\nNot Full Access", log.user)
							log.processed_completely = True
							log.count_in_total = False
							log.save()


					# 	elif out_good:
					# 		log.processed_completely = True
					# 		log.count_in_total = True
					# 		log.save()

					# 	else:
					# 		log.processed_completely = True
					# 		log.count_in_total = False
					# 		log.save()







			unprocessed_logs = User_Log.objects.filter(processed_completely=False,
													).order_by('time_in_jitsi_open')

			change_to_60 = False
			if change_to_60:
				get_all = User_Log.objects.all()
				for log in get_all:
					if log.duration_seconds <60:
						log.duration_seconds = 60
						log.save()

			# OLD WAY below

			process = False

			if process:
				for log in logs:
					day = log.date.weekday()
					print(day)
					day_of_week = calendar.day_name[day]
					print(day_of_week)
					log.day_of_week = str(day_of_week)

					if day_of_week in good_days:
						print("Day is good")
						log.day_good = True
						log.save()

					else:
						if log.user.username == "Buddy_Admin" or log.user.username == "Mickie"\
							or log.room.name== "Orientation" or log.user.username == "JinLee"\
							or log.user.username == "Tails" or log.user.username == "Adolfo" \
							or log.user.username == "Yaejin" or log.user.username =="Chrissy":
							log.day_good = True
							log.save()
						else:
							log.day_good = False
							log.save()

					if log.day_good:
						if log.time_in and log.time_out:
							time_difference = log.time_out - log.time_in
							duration_seconds = time_difference.seconds
							log.duration_seconds = duration_seconds	
							log.needs_attention = False					
							log.save()
							print("Good")
						else:
							log.needs_attention = True
							log.save()
							print("Bad")

			process_needs_attention = False
			if process_needs_attention:
				needs_attention_logs = User_Log.objects.filter(needs_attention=True)
				print("Needs attention count", needs_attention_logs.count())
				
				for item in needs_attention_logs:
					print("Item User", item.user)
					item.user.user_sign_ins.explore_user_logs = True
					item.user.user_sign_ins.save()

			process_monday = False

			if process_monday:
				monday_logs = User_Log.objects.filter(day_of_week="Monday")
				for log in monday_logs:
					print("user, role, room", log.user, log.user.role, log.room)
					if log.room.name == "Orientation":
						log.room_good = True
						log.save()












			# day_good = User_Log.objects.filter(day_good=True).count()

			# # for log in day_good:

			# context['day_is_good_count'] = day_good

			# during_jitsi_count = User_Log.objects.filter(during_jitsi_open=True).count()
			# context['during_jitsi_count'] = during_jitsi_count

			# in_good_count = User_Log.objects.filter(time_in_jitsi_open=True).count()
			# context['in_good_count'] = in_good_count

			# out_good_count = User_Log.objects.filter(time_out_jitsi_open=True).count()
			# context['out_good_count'] = out_good_count

			

			# good_missing_time_count = User_Log.objects.filter(day_good=True, time_out = None).count()
			# context['good_missing_time_count'] = good_missing_time_count

			processed_logs_count = User_Log.objects.filter(processed_completely=True)
			context['processed_logs_count'] = processed_logs_count.count()

			context['display_logs'] = logs		
			return render(request, "site_admin/reports/all_logs.html", context)
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')	

		else:
			return redirect('access_denied')

def individual_time_logs_view(request, user_id):
	context = {}
	context['page_title'] = "User Logs"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			member = CustomUser.objects.get(id=user_id)
			context['member']=member	

			semester = Semester.objects.get(active_semester=True)
			context['semester'] = semester

			

			logs = User_Log.objects.filter(user=member, count_in_total=True).order_by('id')
			context['logs'] = logs

			final = Sign_In_By_User.objects.get(user=member)
			context['final_stats'] = final


			return render(request, "site_admin/reports/individual_logs.html", context)
		else:
			return redirect('access_denied')





def process_all_sign_ins_view(request, role_name):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.is_superuser:
			print("\n\n\nProcess process_all_sign_ins_view", role_name)
			if role_name == "All":
				qs = CustomUser.objects.all().order_by("user_sign_ins__diff_count")
			else:
				role = Role.objects.get(name=role_name)
				qs = CustomUser.objects.filter(role=role).order_by("user_sign_ins__diff_count")

			print("TOTAL", qs.count())
			process = True
			if process:

				for u in qs:
					sign_in_user, created = Sign_In_By_User.objects.get_or_create(user=u)
					all_logs  = User_Log.objects.filter(user=u)
					good_day_logs = User_Log.objects.filter(user=u, count_in_total=True)

					for log in all_logs:
						sign_in_user.logs.add(log)

					for log in good_day_logs:
						sign_in_user.good_day_logs.add(log)


					sign_in_user.diff_count = all_logs.count() - good_day_logs.count()

					problem = A_Problem_User.objects.filter(user=u)
					if problem.count() > 0:
						sign_in_user.problem_user = True

					sign_in_user.save()
					sign_in_user.calculate_total_minutes()
						
			return redirect('site_admin:all_sign_ins', role_name=role_name)
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')	

		else:
			return redirect('access_denied')

def users_signs_ins_view(request, role_name, **kwargs):
	context = {}
	context['page_title'] = "Users Sign Ins"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			if request.GET:
				query = request.GET.get('q', '')
				type_user = request.GET.get('type_user', '')
				if query !='':
					context['users'] = get_users_queryset(query)
					context['count'] = len(context['users'])
					context['role_name'] = "Matching"
				else:
					context['users'] = all_users().order_by('username')
					context['count'] = context['users'].count
					context['role_name'] = "All"

						
			else:
				if role_name == "All":
					context['users'] = CustomUser.objects.all().order_by('username')
					context['count'] = context['users'].count
					context['role_name'] = role_name
				else:
					context['users'] = CustomUser.objects.filter(role__name=role_name).order_by('username')
					context['count'] = context['users'].count
					context['role_name'] = role_name

			
			context = button_bar_context_additions(context)	
			return render(request, "site_admin/reports/total_time.html", context)
		else:
			return redirect('access_denied')

def download_assessments_report(request):
	print("Downloading Assesment Report")
	students = CustomUser.objects.filter(role__name="Student").order_by('username')
	all_count=students.count()
	active_count = students.filter(user_dropped=False).count()	
	dropped_count= students.filter(user_dropped=True).count()
	without_assessments = Student_Progress.objects.filter(user__user_dropped=False, initial_assessment=False).count()
			
	wb = Workbook()

	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	late_color = "faf6b4"
	no_show_color = "edb277" 

	ws.title = "Assessment Report"
	ws['A1'] = "Students"
	ws['A2'] = "Total = " + str(all_count)

	ws['C1'] = "Active"
	ws['C2'] = "Total = " + str(active_count)

	ws['E1'] = "Dropped"
	ws['E2'] = "Total = " + str(dropped_count)

	ws['G1'] = "Needs Assessment"
	ws['G2'] = "Total = " + str(without_assessments)

	ws.append([])
	ws.append(['Username', 'Full Name', 'Needs Assessment', 'Dropped',
				 'Initial', 'Current', 'Final', 'Total', 'Last Assessed'])
	for cell in  ws[4]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=double)
	starting_row = 5
	student_starting_row = 5
	# for row in students:
	# Write data to cell
	for student in students:
		print("Student Row: ", student_starting_row)

		initial = student.student_progress.initial_assessment
		if initial:
			no_initial = ""
		else:
			no_initial = "Yes"

		user_dropped = student.user_dropped
		if user_dropped:
			user_dropped = "Yes"
		else:
			user_dropped = ""

		if student.student_progress.starting:
			starting = student.student_progress.starting.name
		else:
			starting = "None"

		if student.student_progress.current:
			current = student.student_progress.current.name
		else:
			current = "None"

		if student.student_progress.end:
			final = student.student_progress.end.name
		else:
			final = "None"

		a_count = str(student.student_progress.assessments.count())

		if student.student_progress.last_assessed:		
			last_year = student.student_progress.last_assessed.year
			last_month = student.student_progress.last_assessed.month
			last_day=student.student_progress.last_assessed.day
			last_assessed = str(last_month) +"/" + str(last_day) +"/" + str(last_year) 
		else:
			last_assessed="None"


		if not student.student_progress.initial_assessment and not student.user_dropped:
			ws.append([student.username, student.full_name, no_initial, user_dropped,
				 starting, current, final, a_count, last_assessed])
			
		
			for cell in  ws[student_starting_row]:
				cell.fill = PatternFill("solid", fgColor="fcb1b1")

		
			for cell in ws[student_starting_row]:
					cell.border = Border(bottom=thin)

			student_starting_row = student_starting_row + 1

		else:

			ws.append([student.username, student.full_name, no_initial, user_dropped,
				 starting, current, final, a_count, last_assessed])

				
			for cell in ws[student_starting_row]:
					cell.border = Border(bottom=thin)

			student_starting_row = student_starting_row + 1

	# Save the workbook
	response = HttpResponse(content_type='application/ms-excel')

	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=assessment_report_' + str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	response['Content-Disposition'] = file_str

	dims = {}
	for row in ws.rows:
		for cell in row:
			# print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
				if cell.value == "Late":
					# for cell in  ws[student_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=late_color)
					
				elif cell.value == "No Show":
					# for cell in  ws[student_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=no_show_color)
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5
		


	wb.save(response)
	return response


def assessments_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Student Assessments"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			all_students = CustomUser.objects.filter(role__name="Student").order_by('username')
			all_count=all_students.count()
			context['all_students']= all_students
			context['all_count'] = all_count
			context['active_count']=all_students.filter(user_dropped=False).count()	
			context['dropped_count']=all_students.filter(user_dropped=True).count()

			without_assessments = Student_Progress.objects.filter(user__user_dropped=False, initial_assessment=False).count()
			context['without_assessment'] = without_assessments
			return render(request, "site_admin/assessments/assessments.html", context)
		else:
			return redirect('access_denied')

def create_avatar(user):

	if user.role.name == "Staff":
		bg_color = (171, 189, 192)
		text_color = "black"

	elif user.role.name == "Student":
		bg_color=(24, 153, 104)
		text_color = "black"

	elif user.role.name == "Volunteer":
		bg_color = (30, 129, 176)
		text_color = "black"

	print("\n\n\nCreating avatar", user)

	if user.avatar_img:
		user.avatar_img.delete()
		user.save()

	font_path = "static/fonts/consolab.ttf"
	W, H = (200,200)
	img = Image.new('RGBA', (W,H), color = bg_color)
	d = ImageDraw.Draw(img)
	font = ImageFont.truetype(font_path, size=120)
	letters = user.first_name[0].upper() + user.last_name[0].upper()
	w,h = font.getsize(letters)
	d.text(((W-w)/2,(H-h)/2.2), letters, fill=text_color, font=font, stroke_width=1, stroke_fill=(78, 79, 79))

	ava_io = BytesIO() 
	file_name = user.username +"-image.png" 

	img.save(ava_io, 'png', quality=85) # save image to BytesIO object

	avatar = File(ava_io, file_name)
	user.avatar_img = avatar
	user.save()


def all_student_assessments_view(request, **kwargs):
	context = {}
	context['page_title'] = "Student Assessments"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":			
			context = button_bar_context_additions(context)	


			return render(request, "site_admin/reading_levels/all_assessments.html", context)
		else:
			return redirect('access_denied')

def process_session_evaluations_view(request):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.is_superuser:
			print("\n\n\nProcess Session Evaluations")
			evals = End_Session_Evaluation.objects.all()			

			for item in evals:
				# print("\n\nITEM", item)
				if item.level_assessment_performed:
					
					# print("Student", item.get_student())
					if item.get_student():
						
						if item.assessment_level:
							assessment, created = Student_Assessment.objects.get_or_create(
																	user=item.get_student(),
																	date=item.date,
																	assessed_by=item.completed_by,
																	level=item.assessment_level
																	)
							item.get_student().student_progress.assessments.add(assessment)							

						else:
							reason = "Assessment Performed, Student Choosen, No Level"
							incomplete, created = Incomplete_Evaluation.objects.get_or_create(
													evaluation=item,
													user=item.completed_by,
													reason=reason)
							print("\n\n\nYes Assessment")
							print("Student", item.get_student())						
							print("Date", item.date)
							print("No Level Added", item.id)
					else:
						print("\n\n\nNO STUDENT", item.id)
						if item.assessment_level:
							reason = "Assessment Performed, No Student, Has Level"
							incomplete, created = Incomplete_Evaluation.objects.get_or_create(
													evaluation=item,
													user=item.completed_by,
													reason=reason)
						else:
							reason = "Assessment Performed, No Student, No Level"
							incomplete, created = Incomplete_Evaluation.objects.get_or_create(
													evaluation=item,
													user=item.completed_by,
													reason=reason)

				# else:
				# 	print("No Assessment")
				# 	print("Student", item.get_student())

			students = CustomUser.objects.filter(role__name="Student")
			for student in students:

				count = student.student_progress.assessments.all().count()
				print(student, count)
				all_assessments = student.student_progress.assessments.all().order_by('date')
				if count > 0:
					student.student_progress.initial_assessment = True
					student.student_progress.starting = all_assessments.first().level
					student.student_progress.current = all_assessments.last().level
					student.student_progress.last_assessed = all_assessments.last().date
					student.student_progress.save()



						
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')	

		else:
			return redirect('access_denied')

def download_volunteer_report(request):

	print("Downloading Volunteer Report")
	volunteers = CustomUser.objects.filter(role__name="Volunteer").exclude(username__contains="Volunteer").order_by('username')
	total_days_with_sessions = Day_With_Daily_Session.objects.all()
	all_complete= Day_With_Daily_Session.objects.filter(all_complete=True)

	all_sessions = Daily_Session.objects.all()
	all_complete_session= Daily_Session.objects.filter(session_complete=True)
			
	wb = Workbook()


	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	late_color = "faf6b4"
	no_show_color = "eb3528" 

	ws.title = "Volunteer Report"
	ws['A1'] = "Volunteers"
	ws['A2'] = "Total = " + str(volunteers.count())

	ws['E1'] = "Days With Sessions"
	ws['E2'] = "Complete: " + str(all_complete.count()) +" / Total: " + str(total_days_with_sessions.count())

	ws['I1'] = "Sessions"
	ws['I2'] = "Complete: " + str(all_complete_session.count()) +" / Total: " + str(all_sessions.count())

	ws.append([])
	ws.append(['Username', 'Full Name', 'Mega', 'Team', 'No Logins',
				 'Has Current Active Match', 'Days', 'Session Slot', 'Scheduled',
				 'Status', 'First In/Last Out',
				'Total Minutes', 'Temp_Match_Created', 
				 'Assigned Buddies', 'Total Temp Matches',])
	for cell in  ws[4]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=double)
	starting_row = 5
	vol_starting_row = 5
	
	# Write data to cell
	for volunteer in volunteers:
		print("\nVolunteer", volunteer)
		if volunteer.volunteer_profile.mega:
			mega = volunteer.volunteer_profile.mega.name
		else:
			mega = ""

		if volunteer.volunteer_profile.team:
			team = volunteer.volunteer_profile.team.leader.full_name
		else:
			team = ""
		print("VOL Row: ", vol_starting_row)
		temp_buddies = volunteer.volunteer_attendance_report.temporary_buddies.all()
		buddies_string = ""
		# list_buddies = []
		for buddy in temp_buddies:
			if buddy != temp_buddies.last():
				buddies_string = buddies_string + buddy.full_name + " - "
			else:
				buddies_string = buddies_string + buddy.full_name


		no_logs = volunteer.volunteer_attendance_report.no_logs
		if no_logs:
			no_logs = "X"
		else:
			no_logs = ""

		has_active_match = volunteer.volunteer_attendance_report.has_active_match
		if has_active_match:
			has_active_match = "Yes"
		else:
			has_active_match = "No"

		count = 1
		sign_ins = volunteer.sign_ins.all()



		for sign_in in volunteer.sign_ins.all():			
			day_str = str(count) + ". " + sign_in.day.short_day_name()

			if sign_in.session_slot:
				session_slot_str = sign_in.session_slot.session_slot
			else:
				session_slot_str = ""

			if sign_in.scheduled:
				scheduled_str = "Yes"
			else:
				scheduled_str = "No"

			if sign_in.status:
				status_str = sign_in.status.name
			else:
				status_str = ""

			if sign_in.logs.all().count() > 0:

				first_in_time = sign_in.logs.all().first().local_time_in_only_str()
				last_out_time = sign_in.logs.all().last().local_time_out_only_str()


				time_rest = first_in_time + " - " + last_out_time
			else:
				time_rest ="-"

			temp_match_created = sign_in.temp_match_created
			if temp_match_created:
				temp_match_created = "Yes"
			else:
				temp_match_created = "No"

			if sign_in == volunteer.sign_ins.all().first():
				ws.append([volunteer.username, volunteer.full_name, mega, team, no_logs,
					 has_active_match, day_str , session_slot_str, scheduled_str,
					 status_str, time_rest,
					sign_in.total_minutes, temp_match_created, 
					 buddies_string, volunteer.volunteer_attendance_report.total_temp_matches])
			else:
				ws.append(["", "", "", "", "",
					 "", day_str, session_slot_str, scheduled_str,
					 status_str, time_rest,
					sign_in.total_minutes, temp_match_created, 
					 '', '',])

			if no_logs:
				for cell in  ws[vol_starting_row]:
					if cell.value != "Late" or cell.value !="No Show":
						cell.fill = PatternFill("solid", fgColor="fcb1b1")

			if sign_in == volunteer.sign_ins.all().last():
				for cell in ws[vol_starting_row]:
					cell.border = Border(bottom=thin)

			vol_starting_row = vol_starting_row + 1
		
		

	# Save the workbook
	response = HttpResponse(content_type='application/ms-excel')

	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=volunteer_report_' + str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	response['Content-Disposition'] = file_str

	dims = {}
	for row in ws.rows:
		for cell in row:
			# print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
				if cell.value == "Late":
					# for cell in  ws[vol_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=late_color)
					
				elif cell.value == "No Show":
					# for cell in  ws[vol_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=no_show_color)
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5
		


	wb.save(response)
	return response

def download_volunteer_report2(request):

	print("Downloading Volunteer Report")
	volunteers = CustomUser.objects.filter(role__name="Volunteer").exclude(username__contains="Volunteer").order_by('username')
	total_days_with_sessions = Day_With_Daily_Session.objects.all()
	all_complete= Day_With_Daily_Session.objects.filter(all_complete=True)

	all_sessions = Daily_Session.objects.all()
	all_complete_session= Daily_Session.objects.filter(session_complete=True)
			
	wb = Workbook()


	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	late_color = "faf6b4"
	no_show_color = "edb277" 

	ws.title = "Volunteer Report"
	ws['A1'] = "Volunteers"
	ws['A2'] = "Total = " + str(volunteers.count())

	ws['E1'] = "Days With Sessions"
	ws['E2'] = "Complete: " + str(all_complete.count()) +" / Total: " + str(total_days_with_sessions.count())

	ws['I1'] = "Sessions"
	ws['I2'] = "Complete: " + str(all_complete_session.count()) +" / Total: " + str(all_sessions.count())

	ws.append([])
	ws.append(['Username', 'Full Name', 'Mega', 'Team', 'No Logins', 'Has Current Active Match', 'Days', 'Status', 'First In/Last Out',
				'Total Minutes', 'Temp_Match_Created', 'Total Temp Matches', 'Assigned Buddies' ])
	for cell in  ws[4]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=double)
	starting_row = 5
	vol_starting_row = 5
	
	# Write data to cell
	for volunteer in volunteers:
		print("\nVolunteer", volunteer)
		if volunteer.volunteer_profile.mega:
			mega = volunteer.volunteer_profile.mega.name
		else:
			mega = ""

		if volunteer.volunteer_profile.team:
			team = volunteer.volunteer_profile.team.leader.full_name
		else:
			team = ""
		print("VOL Row: ", vol_starting_row)
		temp_buddies = volunteer.volunteer_attendance_report.temporary_buddies.all()
		buddies_string = ""
		# list_buddies = []
		for buddy in temp_buddies:
			if buddy != temp_buddies.last():
				buddies_string = buddies_string + buddy.full_name + " - "
			else:
				buddies_string = buddies_string + buddy.full_name


		no_logs = volunteer.volunteer_attendance_report.no_logs
		if no_logs:
			no_logs = "X"
		else:
			no_logs = ""

		has_active_match = volunteer.volunteer_attendance_report.has_active_match
		if has_active_match:
			has_active_match = "Yes"
		else:
			has_active_match = "No"

		if no_logs:
			ws.append([volunteer.username, volunteer.full_name, mega, team,
			 no_logs, has_active_match,
				 '', '', '',
				'', '', '', '' ])
		# if student.student_attendance_report.no_logs:
			for cell in  ws[vol_starting_row]:
				cell.fill = PatternFill("solid", fgColor="fcb1b1")

		
			for cell in ws[vol_starting_row]:
					cell.border = Border(bottom=thin)

			vol_starting_row = vol_starting_row + 1

		else:
			count = 1
			sign_ins = volunteer.sign_ins.all()
			print("Count of day sign_ins_day", sign_ins.count())


			first_sign_in = sign_ins.first()
			if first_sign_in.session:
				status = first_sign_in.status.name
				
					
				first_day_str = str(count) + '. ' + first_sign_in.session.short_date_day_session_slot()
				
				first_status = first_sign_in.status
				if first_status:
					first_status = first_status.name
				else:
					first_status = ""

				if first_sign_in.logs.all().count() > 0:

					first_in_time = first_sign_in.logs.all().first().local_time_in_only_str()
					last_out_time = first_sign_in.logs.all().last().local_time_out_only_str()


					time_first = first_in_time + " - " + last_out_time
				else:
					time_first ="-"

				temp_match_created = first_sign_in.temp_match_created
				if temp_match_created:
					temp_match_created = "Yes"
				else:
					temp_match_created = "No"

				count = count + 1


				ws.append([volunteer.username, volunteer.full_name, mega, team,
					 no_logs, has_active_match,
					 first_day_str, first_status, time_first, first_sign_in.total_minutes,
					temp_match_created, volunteer.volunteer_attendance_report.total_temp_matches, buddies_string ])
			
			else:
				ws.append([volunteer.username, volunteer.full_name, mega, team,
					 no_logs, has_active_match,
					 "first_day_str", "first_status", "time_first", first_sign_in.total_minutes,
					"temp_match_created", volunteer.volunteer_attendance_report.total_temp_matches, buddies_string ])
			

			
			for sign_in in volunteer.sign_ins.all():
				if sign_in != first_sign_in:
				
					if sign_in.session:
						status = sign_in.status.name
						if sign_in.logs.all().count() > 0:

							first_in_time = sign_in.logs.all().first().local_time_in_only_str()
							last_out_time = sign_in.logs.all().last().local_time_out_only_str()


							time_rest = first_in_time + " - " + last_out_time
						else:
							time_rest ="-"

						temp_match_created = sign_in.temp_match_created
						if temp_match_created:
							temp_match_created = "Yes"
						else:
							temp_match_created = "No"

						ws.append(["", "", "", "", "", "",
						 str(count) + '. ' + sign_in.session.short_date_day_session_slot(), sign_in.status.name, time_rest,
						sign_in.total_minutes, temp_match_created, '', '' ])

						# if status == "Late":
						# 	for cell in  ws[vol_starting_row +1]:
						# 		cell.fill = PatternFill("solid", fgColor=late_color)
								
						# elif status == "No Show":
						# 	for cell in  ws[vol_starting_row +1]:
						# 		cell.fill = PatternFill("solid", fgColor=no_show_color)
							
						vol_starting_row = vol_starting_row + 1
						count = count + 1
			
			
			for cell in ws[vol_starting_row]:
					cell.border = Border(bottom=thin)

			vol_starting_row = vol_starting_row + 1

	# Save the workbook
	response = HttpResponse(content_type='application/ms-excel')

	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=volunteer_report_' + str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	response['Content-Disposition'] = file_str

	dims = {}
	for row in ws.rows:
		for cell in row:
			# print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
				if cell.value == "Late":
					# for cell in  ws[vol_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=late_color)
					
				elif cell.value == "No Show":
					# for cell in  ws[vol_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=no_show_color)
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5
		


	wb.save(response)
	return response

def download_student_report(request):
	print("Downloading Student Report")
	students = CustomUser.objects.filter(role__name="Student", user_dropped=False).exclude(username__contains="Student").order_by('username')
	total_days_with_sessions = Day_With_Daily_Session.objects.all()
	all_complete= Day_With_Daily_Session.objects.filter(all_complete=True)

	all_sessions = Daily_Session.objects.all()
	all_complete_session= Daily_Session.objects.filter(session_complete=True)
			
	wb = Workbook()

	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")
	late_color = "faf6b4"
	no_show_color = "edb277" 

	ws.title = "Student Report"
	ws['A1'] = "Students"
	ws['A2'] = "Total = " + str(students.count())

	ws['E1'] = "Days With Sessions"
	ws['E2'] = "Complete: " + str(all_complete.count()) +" / Total: " + str(total_days_with_sessions.count())

	ws['I1'] = "Sessions"
	ws['I2'] = "Complete: " + str(all_complete_session.count()) +" / Total: " + str(all_sessions.count())

	ws.append([])
	ws.append(['Username', 'Full Name', 'No Logins', 'Has Current Active Match', 'Days', 'Status', 'First In/Last Out',
				'Total Minutes', 'Temp_Match_Created', 'Total Temp Matches', 'Assigned Buddies' ])
	for cell in  ws[4]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=double)
	starting_row = 5
	student_starting_row = 5
	# for row in students:
	# Write data to cell
	for student in students:
		print("Student Row: ", student_starting_row)
		temp_buddies = student.student_attendance_report.temporary_buddies.all()
		buddies_string = ""
		# list_buddies = []
		for buddy in temp_buddies:
			if buddy != temp_buddies.last():
				buddies_string = buddies_string + buddy.full_name + " - "
			else:
				buddies_string = buddies_string + buddy.full_name
			# list_buddies.append(buddy.full_name)

		# if len(list_buddies)==0:
		# 	print("Zero Buddies")
		# 	buddies = ""
		# 	remaining= ""
		# elif len(list_buddies)== 1:
		# 	buddies = list_buddies[0] + " Just One"
		# 	list_buddies.pop(0)
		# 	remaining = "Remaining" + str(len(list_buddies))
		# else:
		# 	buddies = list_buddies[0] + " Multiple"
		# 	remaining = "more than 1 " + str(len(list_buddies))

		no_logs = student.student_attendance_report.no_logs
		if no_logs:
			no_logs = "X"
		else:
			no_logs = ""

		has_active_match = student.student_attendance_report.has_active_match
		if has_active_match:
			has_active_match = "Yes"
		else:
			has_active_match = "No"

		if no_logs:
			ws.append([student.username, student.full_name, no_logs, has_active_match,
				 '', '', '',
				'', '', '', '' ])
		# if student.student_attendance_report.no_logs:
			for cell in  ws[student_starting_row]:
				cell.fill = PatternFill("solid", fgColor="fcb1b1")

		
			for cell in ws[student_starting_row]:
					cell.border = Border(bottom=thin)

			student_starting_row = student_starting_row + 1

		else:
			count = 1
			sign_ins = student.sign_ins.all()


			first_sign_in = sign_ins.first()
			if first_sign_in.session:
				status = first_sign_in.status.name
				
					
				first_day_str = str(count) + '. ' + first_sign_in.session.short_date_day_session_slot()
				
				first_status = first_sign_in.status
				if first_status:
					first_status = first_status.name
				else:
					first_status = ""

				if first_sign_in.logs.all().count() > 0:

					first_in_time = first_sign_in.logs.all().first().local_time_in_only_str()
					last_out_time = first_sign_in.logs.all().last().local_time_out_only_str()


					time_first = first_in_time + " - " + last_out_time
				else:
					time_first ="-"

				temp_match_created = first_sign_in.temp_match_created
				if temp_match_created:
					temp_match_created = "Yes"
				else:
					temp_match_created = "No"

				count = count + 1

			ws.append([student.username, student.full_name, no_logs, has_active_match,
					 first_day_str, first_status, time_first, first_sign_in.total_minutes,
					temp_match_created, student.student_attendance_report.total_temp_matches, buddies_string ])
			
			
			
			for sign_in in student.sign_ins.all():
				if sign_in != first_sign_in:
				
					if sign_in.session:
						status = sign_in.status.name
						if sign_in.logs.all().count() > 0:

							first_in_time = sign_in.logs.all().first().local_time_in_only_str()
							last_out_time = sign_in.logs.all().last().local_time_out_only_str()


							time_rest = first_in_time + " - " + last_out_time
						else:
							time_rest ="-"

						temp_match_created = sign_in.temp_match_created
						if temp_match_created:
							temp_match_created = "Yes"
						else:
							temp_match_created = "No"

						ws.append(["", "", "", "",
						 str(count) + '. ' + sign_in.session.short_date_day_session_slot(), sign_in.status.name, time_rest,
						sign_in.total_minutes, temp_match_created, '', '' ])

						# if status == "Late":
						# 	for cell in  ws[student_starting_row +1]:
						# 		cell.fill = PatternFill("solid", fgColor=late_color)
								
						# elif status == "No Show":
						# 	for cell in  ws[student_starting_row +1]:
						# 		cell.fill = PatternFill("solid", fgColor=no_show_color)
							
						student_starting_row = student_starting_row + 1
						count = count + 1
			
			
			for cell in ws[student_starting_row]:
					cell.border = Border(bottom=thin)

			student_starting_row = student_starting_row + 1

	# Save the workbook
	response = HttpResponse(content_type='application/ms-excel')

	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=student_report_' + str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	response['Content-Disposition'] = file_str

	dims = {}
	for row in ws.rows:
		for cell in row:
			# print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
				if cell.value == "Late":
					# for cell in  ws[student_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=late_color)
					
				elif cell.value == "No Show":
					# for cell in  ws[student_starting_row +1]:
					cell.fill = PatternFill("solid", fgColor=no_show_color)
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5
		


	wb.save(response)
	return response



def download_evaluation_report(request):
	print("Downloading Evaluation Report")
	evals = End_Session_Evaluation.objects.all().order_by('date')
			
	wb = Workbook()

	# Get first sheet
	ws = wb.active
	thin = Side(border_style="thin", color="000000")
	double = Side(border_style="double", color="000000")

	ws.append(['Date', 'Completed By', 'Read With Scheduled', 'Temp Student Assigned',
			'For Student', 'Scheduled Student Attendance', 'Level Assessment Performed',
			'Level From Assessment', 'Books Read', 'Level Of Books Read',
			'Engagement', 'Word Recognition', 'Pronunciation / Fluency',
			'Vocabulary', 'Comprehension', 'Session Comment',
			'Social Emotional Learning Comment', 'Needs Follow Up', 'Follow Up Type',
			 'Follow Up Comment' ])
	for cell in  ws[1]:
		cell.font = Font(bold=True)
		cell.border = Border(bottom=double)

	
	for eval in evals:
		student = eval.get_student()
		if student:
			student_name = eval.get_student().full_name
		else:
			student_name = ""

		scheduled_student_attendance = eval.scheduled_student_attendance
		if scheduled_student_attendance:
			scheduled_student_attendance = scheduled_student_attendance.name
		else:
			scheduled_student_attendance = ""

		assessment_level = eval.assessment_level
		if assessment_level:
			assessment_level = assessment_level.name
		else:
			assessment_level = ""

		engagement = eval.engagement
		if engagement:
			engagement = engagement.name
		else:
			engagement = ""

		word_recognition = eval.word_recognition
		if word_recognition:
			word_recognition = word_recognition.name
		else:
			word_recognition = ""

		pronunciation_fluency = eval.pronunciation_fluency
		if pronunciation_fluency:
			pronunciation_fluency = pronunciation_fluency.name
		else:
			pronunciation_fluency = ""

		vocabulary = eval.vocabulary
		if vocabulary:
			vocabulary = vocabulary.name
		else:
			vocabulary = ""

		comprehension = eval.comprehension
		if comprehension:
			comprehension = comprehension.name
		else:
			comprehension = ""

		follow_up_type = eval.follow_up_type
		if follow_up_type:
			follow_up_type = follow_up_type.name
		else:
			follow_up_type = ""

		ws.append([eval.date, eval.completed_by.full_name, eval.read_with_scheduled,
			eval.temp_student_assigned, student_name, scheduled_student_attendance,
			eval.level_assessment_performed, assessment_level, eval.books_read,
			eval.get_level_read_str(), engagement, word_recognition, 
			pronunciation_fluency, vocabulary, comprehension, eval.session_comment,
			eval.social_emotional_learning_comment, eval.follow_up_needed,
			follow_up_type, eval.follow_up_comment])
	date=timezone.localtime(timezone.now())
	file_str = 'attachment; filename=evaluations_report_' + str(date.month) + '-' + str(date.day) + '-' + str(date.year) +'.xlsx'
	

	dims = {}
	for row in ws.rows:
		for cell in row:
			print(cell, cell.value)
			# cell_value = str(cell.value)
			# if cell_value != "":
			if cell.value:
				dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
			   
	for col, value in dims.items():
		ws.column_dimensions[col].width = value + 5


	response = HttpResponse(content_type='application/ms-excel')

	# Set the content disposition as attachment and provide the filename
	# response['Content-Disposition'] = 'attachment; filename=evaluations_report.xlsx'
	response['Content-Disposition'] = file_str

	wb.save(response)
	return response

def process_match_attendance_records(day):
	for session in day.day_sessions.all():
		print("\n\n\n*******************Session", session)
		temporary_matches = Temporary_Match.objects.filter(session=session)
		print("\n\nTemporary Matches", temporary_matches)
		reassigned_students = []
		reassigned_volunteers = []
		for match in temporary_matches:
			print("\n\nMatch", match)
			reassigned_students.append(match.student_user)
			# 
			reassigned_volunteers.append(match.teacher_user)
			# 

		print("reassigned_students", reassigned_students)
		print("reassigned_volunteers", reassigned_volunteers)

		for item in reassigned_students:
			all_match_attendance_records_stu = Match_Attendance_Record.objects.filter(
															session=session,
															sch_match__student = item)
			print("\n\n\nall_match_attendance_records_stu count", all_match_attendance_records_stu.count())
			for item in all_match_attendance_records_stu:
				print("record", item.id, item,  )
				item.member_reassigned = True
				item.student_reassigned = True
				item.save()

		for item in reassigned_volunteers:
			all_match_attendance_records_vol= Match_Attendance_Record.objects.filter(
															session=session,
															sch_match__volunteer= item)
			print("\n\n\nall_match_attendance_records_vol count", all_match_attendance_records_vol.count())
			for item in all_match_attendance_records_vol:
				print("record", item.id, item,  )
				item.member_reassigned = True
				item.volunteer_reassigned = True
				item.save()


		all_match_attendance_records = Match_Attendance_Record.objects.filter(session=session)
		print("all_match_attendance_records", all_match_attendance_records.count())

		for record in all_match_attendance_records:
			student_in_record = record.get_student()
			student_user_logs = User_Log.objects.filter(user=student_in_record, date=session.date)
			print("Student User logs", student_in_record, student_user_logs.count())
			if student_user_logs.count() > 0:
				record.student_present = True

			buddy_in_record = record.get_buddy()
			buddy_user_logs = User_Log.objects.filter(user=buddy_in_record, date=session.date)
			print("Buddy User logs", buddy_in_record, buddy_user_logs.count())
			if buddy_user_logs.count() > 0:
				record.volunteer_present = True

			record.save()



			# participants_sch_matches = Scheduled_Match.objects.filter(
			# 							Q(student = match.student_user)|
			# 							Q(volunteer = match.teacher_user))

			# print("participants_sch_matches", participants_sch_matches)
			# for sch_match in participants_sch_matches:
			# 	print("\n\nScheduled Match", sch_match)
			# 	match_records = Match_Attendance_Record.objects.filter(session=session,
			# 														match_type__name="Scheduled",
			# 														sch_match=sch_match)
				# print("match record count",sch_match, match_records.count())
				# for sch_match_record in match_attendance_records
				# print("Match Record", match_record, match_record.id)
				# match_record.member_reassigned = True
				# if sch_match.student in reassigned_students:
				# 	match_record.student_reassigned = True
				# if sch_match.volunteer in reassigned_volunteers:
				# 	match_record.volunteer_reassigned = True
				# match_record.save()
				# print("Match Record After")
				# print("MR", match_record.member_reassigned )
				# print("SR", match_record.student_reassigned )
				# print("VR", match_record.volunteer_reassigned)

def attendance_by_session_view(request, session_id):
	context = {}
	context['page_title'] = "Attendance By Session"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			session = Daily_Session.objects.get(id = session_id)
			semester = session.semester
			context['session'] = session
			context['semester'] = semester
			return render(request, "site_admin/reports/attendance_records_by_session.html", context)
		else:
			return redirect('access_denied')


def attendance_by_day_view(request, day_id):
	context = {}
	context['page_title'] = "Attendance By Day"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			return render(request, "site_admin/reports/attendance_records_by_day.html", context)
		else:
			return redirect('access_denied')




def get_volunteer_queryset(query=None):
	queryset = []
	queries = query.split(" ")

	for q in queries:
		users = CustomUser.objects.filter(Q(role__name__icontains="Volunteer"),
			Q(first_name__icontains=q)|
			Q(middle_name__icontains=q)|
			Q(last_name__icontains=q)|
			Q(email__icontains=q)|
			Q(username__icontains=q))

		for user in users:
			queryset.append(user)

	return list(set(queryset))

def update_volunteer_report(vol):
	report, created = Volunteer_Report.objects.get_or_create(user=vol, semester=active_semester())
	user_logs = User_Log.objects.filter(user=vol)

	if user_logs.count() == 0:
		report.no_logs = True
		report.total_sign_ins = user_logs.count()
	else:
		report.no_logs = False
		report.total_sign_ins = user_logs.count()

	if Scheduled_Match.objects.filter(volunteer=vol, match_active=True).exists():
		report.has_active_match = True
	else:
		report.has_active_match = False

	match_attendance_records = Match_Attendance_Record.objects.filter(
		Q(sch_match__volunteer = vol),
		Q(match_type__name ="Scheduled"))

	report.sessions_scheduled.clear()
	for record in match_attendance_records:
		record_session = record.session
		report.sessions_scheduled.add(record_session)

		
	report.save()


def all_volunteers_attendance_report_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "All Volunteers Attendance"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)

			total_days_with_sessions = Day_With_Daily_Session.objects.all()
			all_complete= Day_With_Daily_Session.objects.filter(all_complete=True)
			context['days_with_sessions'] = total_days_with_sessions
			context['all_complete'] = all_complete	

			all_sessions = Daily_Session.objects.all()
			all_complete_session= Daily_Session.objects.filter(session_complete=True)
			context['all_sessions'] = all_sessions
			context['all_complete_session'] = all_complete_session

			volunteers = CustomUser.objects.filter(role__name="Volunteer").exclude(username__contains="Volunteer").order_by('username')

			
			
			for vol in volunteers:
				update_volunteer_report(vol)
				
			
			if request.GET:
				query = request.GET.get('q', '')
				if query !='':
					context['volunteers'] = get_volunteer_queryset(query)
					context['count'] = len(context['volunteers'])
				else:
					context['volunteers'] = volunteers
					context['count'] = context['volunteers'].count

			else:
				
				context['volunteers']=volunteers
				context['count'] = volunteers.count()

			return render(request, "site_admin/reports/all_volunteers_attendance.html", context)
		else:
			return redirect('access_denied')

def individual_volunteer_attendance_report_view(request, user_id):
	context = {}
	context['page_title'] = "Individual Volunteer Attendance"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			volunteer = CustomUser.objects.get(id=user_id)
			context['volunteer']=volunteer	

			semester = Semester.objects.get(active_semester=True)
			context['semester'] = semester

			update_volunteer_report(volunteer)

			report = Volunteer_Report.objects.get(user=volunteer, semester=semester)
			context['report'] = report

			sign_ins_day = Sign_In_By_Day.objects.filter(user=volunteer)
			context['sign_ins_day'] = sign_ins_day

			# days_w_session = Day_With_Daily_Session.objects.all().order_by('date')
			# context['days_w_session'] = days_w_session

			return render(request, "site_admin/reports/individual_volunteer_attendance.html", context)
		else:
			return redirect('access_denied')

def get_student_queryset(query=None):
	queryset = []
	queries = query.split(" ")

	for q in queries:
		users = CustomUser.objects.filter(Q(role__name__icontains="Student"),
			Q(first_name__icontains=q)|
			Q(middle_name__icontains=q)|
			Q(last_name__icontains=q)|
			Q(email__icontains=q)|
			Q(username__icontains=q))

		for user in users:
			queryset.append(user)

	return list(set(queryset))

def update_student_report(stu):
	report, created = Student_Report.objects.get_or_create(user=stu, semester=active_semester())
	user_logs = User_Log.objects.filter(user=stu)

	if user_logs.count() == 0:
		report.no_logs = True
		report.total_sign_ins = user_logs.count()
	else:
		report.no_logs = False
		report.total_sign_ins = user_logs.count()

	if Scheduled_Match.objects.filter(student=stu, match_active=True).exists():
		report.has_active_match = True
	else:
		report.has_active_match = False

	match_attendance_records = Match_Attendance_Record.objects.filter(
		Q(sch_match__student = stu),
		Q(match_type__name ="Scheduled"))

	report.sessions_scheduled.clear()
	for record in match_attendance_records:
		record_session = record.session
		report.sessions_scheduled.add(record_session)

		
	report.save()


def all_students_attendance_report_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "All Students Attendance"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)

			total_days_with_sessions = Day_With_Daily_Session.objects.all()
			all_complete= Day_With_Daily_Session.objects.filter(all_complete=True)
			context['days_with_sessions'] = total_days_with_sessions
			context['all_complete'] = all_complete	

			all_sessions = Daily_Session.objects.all()
			all_complete_session= Daily_Session.objects.filter(session_complete=True)
			context['all_sessions'] = all_sessions
			context['all_complete_session'] = all_complete_session

			all_students = CustomUser.objects.filter(role__name="Student", user_dropped=False).exclude(username__contains="Student").order_by('username')

			
			
			for stu in all_students:
				update_student_report(stu)


			if request.GET:
				query = request.GET.get('q', '')
				if query !='':
					context['students'] = get_student_queryset(query)
					context['count'] = len(context['students'])
				else:
					students = CustomUser.objects.filter(role__name="Student", user_dropped=False).exclude(username__contains="Student").order_by('username')

					context['students'] = students
					context['count'] = context['students'].count

			else:
				# students= CustomUser.objects.filter(role__name="Student", user_dropped=False).order_by('username').exclude(username__contains="Student")
				context['students']=all_students
				context['count'] = all_students.count()

			return render(request, "site_admin/reports/all_students_attendance.html", context)
		else:
			return redirect('access_denied')

def individual_student_attendance_report_view(request, user_id):
	context = {}
	context['page_title'] = "Individual Student Attendance"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			student = CustomUser.objects.get(id=user_id)
			context['student']=student	

			semester = Semester.objects.get(active_semester=True)
			context['semester'] = semester

			update_student_report(student)

			report = Student_Report.objects.get(user=student, semester=semester)
			context['report'] = report

			sign_ins_day = Sign_In_By_Day.objects.filter(user=student)
			context['sign_ins_day'] = sign_ins_day

			# days_w_session = Day_With_Daily_Session.objects.all().order_by('date')
			# context['days_w_session'] = days_w_session

			return render(request, "site_admin/reports/individual_student_attendance.html", context)
		else:
			return redirect('access_denied')

def upload_students(request, *args, **kwargs):
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.is_superuser:
			print("\n\n\nUpload Students")
			
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')
		else:
			return redirect('access_denied')

def upload_volunteers(request, *args, **kwargs):
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.is_superuser:
			print("\n\n\nUpload Volunteers")
			
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')
		else:
			return redirect('access_denied')

def upload_staff(request, *args, **kwargs):
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.is_superuser:
			print("\n\n\nUpload Staff")
			
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')
		else:
			return redirect('access_denied')

# def days_hours_minutes(td):
# 	return td.seconds

def fill_in_time(log, log_list, user):
	if log.time_in and not log.time_out:
		print("\n\nMissing Time Out", user)
		

		for index, item in enumerate(log_list):
			if not item.time_out:
				missing_out_index = index
				next_index_in_list = missing_out_index + 1
				# print("OUT", item.id, item.time_out.date(), item.time_out.time())

		try:
			use_this = None
			for index, item in enumerate(log_list):
				if index == missing_out_index:
					to_fix = item
					print("This one is missing", index, to_fix.id)
				elif index == next_index_in_list:
					print("Use this time IN", index)
					use_this = log_list[next_index_in_list]
					print("use_this", use_this.id, use_this.time_in.date(), use_this.time_in.time())
					to_fix.time_out = use_this.time_in
					to_fix.save()
					time_difference = to_fix.time_out - to_fix.time_in
					duration_seconds = time_difference.seconds
					to_fix.duration_seconds = duration_seconds						
					to_fix.save()
		except Exception as e:
			print("*************No NEXT", e)

def volunteer_sign_ins(day):
	volunteers = CustomUser.objects.filter(role__name="Volunteer").exclude(username__contains="Volunteer")
	for vol in volunteers:
		# if vol.username == "Madelyn":
		day_sign_ins, created = Sign_In_By_Day.objects.get_or_create(user=vol, day=day)
		day_user_logs = User_Log.objects.filter(user=vol, date=day.date)
		
		day_match_attendance_records = Match_Attendance_Record.objects.filter(
			Q(sch_match__volunteer = vol) ,
			Q(session__date =day.date))

		if day_match_attendance_records.count() > 0:
			day_sign_ins.scheduled = True
			if day_match_attendance_records.count() == 1:
				this_record = day_match_attendance_records.first()
				session_slot = this_record.session.day_time
				day_sign_ins.session_slot = session_slot
				day_sign_ins.session = this_record.session
				first_log = day_user_logs.first()
				if day_user_logs.count() > 0:
					time_difference = timezone.localtime(first_log.time_in) - timezone.localtime(this_record.session.session_start_date_time)
					time_diff_sec = time_difference.seconds
					if timezone.localtime(first_log.time_in).time() <= session_slot.time_start:
						day_sign_ins.status = Attendance_Status.objects.get(name="On Time")
						print("on time", time_diff_sec)
					elif timezone.localtime(first_log.time_in).time() >= session_slot.time_start:
						day_sign_ins.status = Attendance_Status.objects.get(name="Late")
						print("late", time_diff_sec)
					# print(day_sign_ins.status)
				else:
					day_sign_ins.status = Attendance_Status.objects.get(name="No Show")
					print("NO Show")
			else:
				print("\n\n COUNT Attendance Record Not = 1", vol, day_match_attendance_records.count())
		else:
			# print("\n\n COUNT Attendance Record 0", vol, day_match_attendance_records.count())
			day_sign_ins.scheduled = False

		day_sign_ins.total_sign_ins = day_user_logs.count()

		if day_user_logs.count() > 0:
			day_sign_ins.signed_in = True

		day_sign_ins.logs.clear()

		previous_log = None	
		log_list = list(day_user_logs)
		missing_out_index = None
		next_index_in_list = None


		for log in day_user_logs:			
			day_sign_ins.logs.add(log)		
			if log.time_in and log.time_out:
				time_difference = log.time_out - log.time_in
				duration_seconds = time_difference.seconds
				log.duration_seconds = duration_seconds						
				log.save()
				previous_log = log
			else:
				if day_user_logs.count() > 1:
					fill_in_time(log, log_list, vol)
				else:
					print("*******************Only one log VOLUNTEER", vol)
					if End_Session_Evaluation.objects.filter(completed_by=vol, date=day.date).exists:
						session_eval = End_Session_Evaluation.objects.filter(completed_by=vol, date=day.date).last()
						log.time_out = session_eval.date_created
						log.save()
					else:
					
						time_in = log.time_in
						for session in day.day_sessions.all():
							print("time_in", time_in)
							print("session.entry_allowed_start", session.entry_allowed_start)
							print("session.entry_allowed_end", session.entry_allowed_end)

							if time_in >= session.entry_allowed_start and time_in <= session.entry_allowed_end:
								log.time_out = session.session_end_date_time
								log.save()
								print("This One", log.time_out)
							else:
								print("NO MATCH for last time")

				

		if day_user_logs.count() > 1:

			tried=0		
			times_try = 5		
			

			still_missing = User_Log.objects.filter(user=vol, date=day.date, time_out=None)	
			print("Still Missing", still_missing)
			# if still_missing.count() > 0:
			# 	for log in still_missing:				
			# 		fill_in_time(log, log_list)	

			while still_missing.count() > 0 and tried <=5:
				for log in still_missing:				
					fill_in_time(log, log_list, vol)
					still_missing = User_Log.objects.filter(user=vol, date=day.date, time_out=None)	
					print("In Loop, Still Missing", still_missing)
					tried = tried + 1		
  		
		final_missing = User_Log.objects.filter(user=vol, date=day.date, time_out=None)	
		print("Final Missing VOLUNTEER", vol, final_missing )

		# day_sign_ins.save()

		day_temp_match_attendance_records = Match_Attendance_Record.objects.filter(
			Q(match_type__name="Temporary"),
			Q(temp_match__teacher_user = vol) ,
			Q(session__date =day.date))

		total_temp_matches =Temporary_Match.objects.filter(teacher_user = vol)
		vol.volunteer_attendance_report.total_temp_matches = total_temp_matches.count()
		vol.volunteer_attendance_report.save()

		print("\n\nVolunteer Temp Match Records", vol)
		print("QS", day_temp_match_attendance_records)

		if day_temp_match_attendance_records.count() > 0:
			print("Volunteer had a temp match")
			day_sign_ins.temp_match_created = True
			
			for record in day_temp_match_attendance_records:
				assigned_buddy = record.get_student()
				print("Temp Student", assigned_buddy)
				
				# vol.volunteer_attendance_report.temporary_buddies.clear()
				vol.volunteer_attendance_report.temporary_buddies.add(assigned_buddy)
				
				# day_sign_ins.temporary_buddies.clear()
				day_sign_ins.temporary_buddies.add(assigned_buddy)
			

		else:
			print("Volunteer had NO temp Match")
			day_sign_ins.temp_match_created = False


		day_sign_ins.calculate_total_minutes()


def student_sign_ins(day):

	students = CustomUser.objects.filter(role__name="Student").exclude(username__contains="Student")
	for stu in students:
		print("\n\n\n\n ############# STUDENT", stu)
		day_sign_ins, created = Sign_In_By_Day.objects.get_or_create(user=stu, day=day)
		day_user_logs = User_Log.objects.filter(user=stu, date=day.date)
		
		day_match_attendance_records = Match_Attendance_Record.objects.filter(
			Q(sch_match__student = stu) ,
			Q(session__date =day.date))

		if day_match_attendance_records.count() > 0:
			# print("\n\n COUNT Attendance Record > 0", stu, day_match_attendance_records.count())
			day_sign_ins.scheduled = True
			if day_match_attendance_records.count() == 1:
				this_record = day_match_attendance_records.first()
				session_slot = this_record.session.day_time
				day_sign_ins.session_slot = session_slot
				day_sign_ins.session = this_record.session
				first_log = day_user_logs.first()
				if day_user_logs.count() > 0:
					time_difference = timezone.localtime(first_log.time_in) - timezone.localtime(this_record.session.session_start_date_time)
					time_diff_sec = time_difference.seconds
					if timezone.localtime(first_log.time_in).time() <= session_slot.time_start:
						day_sign_ins.status = Attendance_Status.objects.get(name="On Time")
						print("on time", time_diff_sec)
					elif timezone.localtime(first_log.time_in).time() >= session_slot.time_start:
						day_sign_ins.status = Attendance_Status.objects.get(name="Late")
						print("late", time_diff_sec)

					# print("FIRST LOG TIME IN", timezone.localtime(first_log.time_in))
					# print("START TIME", timezone.localtime(this_record.session.session_start_date_time))
					# time_difference = timezone.localtime(first_log.time_in) - timezone.localtime(this_record.session.session_start_date_time)
					# time_diff_sec = time_difference.seconds
					# print("Time DIfference FOR STATUS", time_diff_sec)
					# if time_diff_sec < 61:
					# 	day_sign_ins.status = Attendance_Status.objects.get(name="On Time")
					# 	print("ON TIME")
					# else:
					# 	day_sign_ins.status = Attendance_Status.objects.get(name="Late")
					# 	print("Late")

					# print("FIRST LOG TIME IN", timezone.localtime(first_log.time_in))
					# print("START TIME", timezone.localtime(this_record.session.session_start_date_time))
					# time_difference = timezone.localtime(first_log.time_in) - timezone.localtime(this_record.session.session_start_date_time)
					# print("Time DIfference FOR STATUS", time_difference.seconds)
					# if timezone.localtime(first_log.time_in).time() <= session_slot.time_start:
					# 	day_sign_ins.status = Attendance_Status.objects.get(name="On Time")
					# elif timezone.localtime(first_log.time_in).time() >= session_slot.time_start:
					# 	day_sign_ins.status = Attendance_Status.objects.get(name="Late")
					# print(day_sign_ins.status)
				else:
					day_sign_ins.status = Attendance_Status.objects.get(name="No Show")
					print("NO Show")

			else:
				print("COUNT Attendance Record Not = 1", stu, day_match_attendance_records.count())
		else:
			# print("\n\n COUNT Attendance Record 0", stu, day_match_attendance_records.count())
			day_sign_ins.scheduled = False

		day_sign_ins.total_sign_ins = day_user_logs.count()

		if day_user_logs.count() > 0:
			day_sign_ins.signed_in = True

		day_sign_ins.logs.clear()

		previous_log = None	
		log_list = list(day_user_logs)
		missing_out_index = None
		next_index_in_list = None


		for log in day_user_logs:			
			day_sign_ins.logs.add(log)		
			if log.time_in and log.time_out:
				time_difference = log.time_out - log.time_in
				duration_seconds = time_difference.seconds
				log.duration_seconds = duration_seconds						
				log.save()
				previous_log = log
			else:
				
				if day_user_logs.count() > 1:
					fill_in_time(log, log_list, stu)
				else:
					print("*******************Only one log STudent", student)
					time_in = log.time_in
					for session in day.day_sessions.all():
						print("time_in", time_in)
						print("session.entry_allowed_start", session.entry_allowed_start)
						print("session.entry_allowed_end", session.entry_allowed_end)

						if time_in >= session.entry_allowed_start and time_in <= session.entry_allowed_end:
							log.time_out = session.session_end_date_time
							log.save()
							print("This One", log.time_out)
						else:
							print("NO MATCH for last time")
				

				

		if day_user_logs.count() > 1:

			tried=0		
			times_try = 5

			still_missing = User_Log.objects.filter(user=stu, date=day.date, time_out=None)	
			print("Still Missing", still_missing)
			# if still_missing.count() > 0:
			# 	for log in still_missing:				
			# 		fill_in_time(log, log_list)	

			while still_missing.count() > 0 and tried <=5:
				for log in still_missing:				
					fill_in_time(log, log_list, stu)
					still_missing = User_Log.objects.filter(user=stu, date=day.date, time_out=None)	
					print("In Loop, Still Missing", still_missing)
					tried = tried + 1

		final_missing = User_Log.objects.filter(user=stu, date=day.date, time_out=None)	
		print("Final Missing STUDENT", stu, final_missing )	
  		
		

		# day_sign_ins.save()

		day_temp_match_attendance_records = Match_Attendance_Record.objects.filter(
			Q(match_type__name="Temporary"),
			Q(temp_match__student_user = stu) ,
			Q(session__date =day.date))

		total_temp_matches =Temporary_Match.objects.filter(student_user = stu)
		
		stu.student_attendance_report.total_temp_matches = total_temp_matches.count()
		stu.student_attendance_report.save()

		print("\n\nStudent Temp Match Records", stu)
		print("QS", day_temp_match_attendance_records)

		if day_temp_match_attendance_records.count() > 0:
			print("Student had a temp match")
			day_sign_ins.temp_match_created = True
			for record in day_temp_match_attendance_records:
				assigned_buddy = record.temp_match.teacher_user
				stu.student_attendance_report.temporary_buddies.add(assigned_buddy)
				day_sign_ins.temporary_buddies.add(assigned_buddy)
			

		else:
			print("Student had NO temp Match")
			day_sign_ins.temp_match_created = False


		day_sign_ins.calculate_total_minutes()

def adjust_match_attendance_for_student(user):
	pass

def adjust_match_attendance_for_volunteer(user):
	pass


def staff_mark_day_as_complete_view(request, day_id):

	print("\n\n\n\n\n\n***Mark Day as Complete")
	context = {}
	context['page_title'] = "Session Complete Attendance"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			# context['active_semester'] = active_semester()	
			day = Day_With_Daily_Session.objects.get(id = day_id)
			temporary_type = Match_Type.objects.get(name="Temporary")			
			scheduled_type = Match_Type.objects.get(name="Scheduled")

			manual_on_users = User_Session_Status.objects.filter(manual_redirect_on=True)
			for item in manual_on_users:
				item.manual_redirect_on = False
				item.save()


			for session in day.day_sessions.all():		
				session.session_complete = True
				session.save()

				today_temp_matches = Temporary_Match.objects.filter(session=session)
				temp_type = Match_Type.objects.get(name="Temporary")
				for temp_match in today_temp_matches:
					temp_attendance_record, created = Match_Attendance_Record.objects.get_or_create(
						session=session, match_type=temp_type, temp_match=temp_match )
					temp_match.match_active = False
					temp_match.save()
					if temp_match.student_user.session_status.scheduled_match:
						temp_match.student_user.session_status.current_active_match_type = scheduled_type
						temp_match.student_user.session_status.temp_match = None
						temp_match.student_user.session_status.temporary_buddy = None
						temp_match.student_user.session_status.needs_new_buddy = False
					else:
						temp_match.student_user.session_status.current_active_match_type = None
						temp_match.student_user.session_status.temp_match = None
						temp_match.student_user.session_status.temporary_buddy = None
						temp_match.student_user.session_status.needs_new_buddy = False

					temp_match.student_user.session_status.save()

					if temp_match.teacher_user.session_status.scheduled_match:
						temp_match.teacher_user.session_status.current_active_match_type = scheduled_type
						temp_match.teacher_user.session_status.temp_match = None
						temp_match.teacher_user.session_status.temporary_buddy = None
						temp_match.teacher_user.session_status.needs_new_buddy = False
					else:
						temp_match.teacher_user.session_status.current_active_match_type = None
						temp_match.teacher_user.session_status.temp_match = None
						temp_match.teacher_user.session_status.temporary_buddy = None
						temp_match.teacher_user.session_status.needs_new_buddy = False

					temp_match.teacher_user.session_status.save()

					volunteers_original_student = temp_match.teacher_user.session_status.scheduled_buddy
					if volunteers_original_student:
						volunteers_original_student.session_status.needs_new_buddy = False
						volunteers_original_student.session_status.save()

					students_original_volunteer = temp_match.student_user.session_status.scheduled_buddy
					if students_original_volunteer:
						students_original_volunteer.session_status.needs_new_buddy = False
						students_original_volunteer.session_status.save()

			student_sign_ins(day)
			volunteer_sign_ins(day)
			process_match_attendance_records(day)


			# day.all_complete = True	
			# day.save()
			day.set_final_stats()
			return redirect('site_admin:admin_home')
			# return redirect('site_admin:attendance_by_day', day_id=day.id)
		else:
			return redirect('access_denied')

def button_bar_context_additions(context):
	all_roles = Role.objects.all()
	context['all_roles'] = all_roles
	context['active_semester'] = active_semester()
	return context

def create_match_for_semester_view(request, semester_id):
	context = {}
	context['page_title'] = "Create Match"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester				
			student_role = Role.objects.get(name="Student")
			available_students= CustomUser.objects.filter(role=student_role,
														student_profile__match_needed=True)
			context['available_students'] = available_students

			vol_role = Role.objects.get(name="Volunteer")
			available_volunteers= CustomUser.objects.filter(role=vol_role,
															volunteer_profile__match_needed=True)
			context['available_volunteers'] = available_volunteers

			days = semester.days.all()
			context['days'] = days

			day_time_slots = semester.day_time_slots.all().order_by('day__number', 'session_slot')
			context['day_time_slots'] = day_time_slots			

			if request.POST:
				form = Create_Scheduled_Match_Form(request.POST)

				if form.is_valid():							
					match = form.save()

					note_category = Note_Category.objects.get(name="Scheduled Match Created")
					new_note = Note.objects.create(category=note_category, author=request.user,
												   content="New Match Created")
					match.notes.add(new_note)

					#Adjust User Session Status
					scheduled_match_type = Match_Type.objects.get(name="Scheduled")

					student_user_session_status = User_Session_Status.objects.get(user=match.student) 
					student_user_session_status.current_active_match_type = scheduled_match_type
					student_user_session_status.scheduled_match = match
					student_user_session_status.scheduled_buddy = match.volunteer
					student_user_session_status.save()

					volunteer_user_session_status = User_Session_Status.objects.get(user=match.volunteer)
					volunteer_user_session_status.current_active_match_type = scheduled_match_type
					volunteer_user_session_status.scheduled_match = match
					volunteer_user_session_status.scheduled_buddy = match.student
					volunteer_user_session_status.save()

					for slot in match.scheduled_slots.all():
						match.student.student_profile.scheduled_day_time_slots.add(slot)
						match.volunteer.volunteer_profile.scheduled_day_time_slots.add(slot)

					start_date = timezone.now().date()
					end_date = semester.end_date

					semester_daily_sessions = Daily_Session.objects.filter(semester=semester,
									date__range=(start_date, end_date),
									day_time__in=match.scheduled_slots.all())

					both_missing = Match_Status_Option.objects.get(name="Both Missing")

					for session in semester_daily_sessions:						
						# print("\nSession", session)

						if session.date == start_date:
							if session.session_start_date_time > timezone.now():
								# print("Adding")
								match.sessions_scheduled.add(session)
								match_status = Match_Status.objects.create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=match,
															status=both_missing)
								match_attendance_record = Match_Attendance_Record.objects.create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=match)
								session_reading_level = Session_Reading_Level.objects.create(
															session=session,
															user=match.student)
								match.student.student_progress.level_progress.add(session_reading_level)
							# else:
							# 	print("NO ADD")
								
						else:
							match.sessions_scheduled.add(session)
							match_status = Match_Status.objects.create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=match,
															status=both_missing)
							match_attendance_record = Match_Attendance_Record.objects.create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=match)
							session_reading_level = Session_Reading_Level.objects.create(
															session=session,
															user=match.student)
							match.student.student_progress.level_progress.add(session_reading_level)

					messages.success(request, 'Match Created Successfully.')
					return redirect("site_admin:scheduled_matches", semester_id=semester.id,
						by_call= "Semester", by_type="Active")
				else:
					print(form.errors)
					context['form'] = form

			else:
				form = Create_Scheduled_Match_Form()
				context['form'] = form
			
			return render(request, "site_admin/matches/create_match.html", context)
		else:
			return redirect('access_denied')

def edit_match_view(request, match_id):
	context = {}
	context['page_title'] = "Edit Match"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)				
			match = Scheduled_Match.objects.get(id=match_id)
			context['match'] = match
			semester = match.semester
			day_time_slots = semester.day_time_slots.all()
			context['day_time_slots'] = day_time_slots

			days = semester.days.all()
			context['days'] = days

			current_slots = match.scheduled_slots.all()
			text_current = ""
			if current_slots.count() != 0:
				if current_slots.count() == 1:
					slot = current_slots.first()
					text_current = text_current + slot.get_short_name()

				else:
					for slot in current_slots:
						if slot != current_slots.last():
							text_current = text_current + slot.get_short_name() + ", "
						else:
							text_current = text_current + slot.get_short_name()

			else:
				text_current = "No Session Slots"	
			
			current_sessions = match.sessions_scheduled.all()	

			if request.POST:
				form = Edit_Scheduled_Match_Form(request.POST, instance=match)				

				if form.is_valid():
					updated_match = form.save()
					new_slots = updated_match.scheduled_slots.all()
					text_new = ""

					if new_slots.count() != 0:
						if new_slots.count() == 1:
							slot = new_slots.first()
							text_new = text_new + slot.get_short_name()

						else:
							for slot in new_slots:
								if slot != new_slots.last():
									text_new = text_new + slot.get_short_name() + ", "
								else:
									text_new = text_new + slot.get_short_name()

					else:
						text_new = "No Session Slots"	


					note_category= Note_Category.objects.get(name = "Scheduled Match Edit")
					text = text_current + " changed to " + text_new
					new_note = Note.objects.create(category=note_category, author=request.user,
												   content=text)
					updated_match.notes.add(new_note)

					updated_match.student.student_profile.scheduled_day_time_slots.clear()
					updated_match.volunteer.volunteer_profile.scheduled_day_time_slots.clear()

					for slot in new_slots:
						updated_match.student.student_profile.scheduled_day_time_slots.add(slot)
						updated_match.volunteer.volunteer_profile.scheduled_day_time_slots.add(slot)


					start_date = timezone.now().date()
					end_date = semester.end_date

					semester_daily_sessions = Daily_Session.objects.filter(semester=semester,
									date__range=(start_date, end_date),
									day_time__in=match.scheduled_slots.all())

					updated_sessions = []

					for session in semester_daily_sessions:
						updated_sessions.append(session)

					both_missing = Match_Status_Option.objects.get(name="Both Missing")

					for session in current_sessions:
						# print("\n current Session Schedule in Match", session)

						if session.session_start_date_time > timezone.now():
							# print("In the future", session)
							if not session in updated_sessions:
								# print("needs to be removed")
								updated_match.sessions_scheduled.remove(session)
								# print("Delete Match Status and Attendance Record")
								match_status = Match_Status.objects.get(
															session=session,
															sch_match=updated_match)
								match_status.delete()

								match_attendance_record = Match_Attendance_Record.objects.get(
															session=session,
															sch_match=updated_match)
								match_attendance_record.delete()

								session_reading_level = Session_Reading_Level.objects.get(
															session=session,
															user=match.student)
								session_reading_level.delete()

							# else:
							# 	print("Leave it, Still in Session", session)

						# else:
						# 	print("Session is Done", session)

					scheduled_match_type = Match_Type.objects.get(name="Scheduled")

					for session in updated_sessions:
						if session.date == start_date:
							if session.session_start_date_time > timezone.now():
								# print("Adding")
								updated_match.sessions_scheduled.add(session)
								match_status = Match_Status.objects.get_or_create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=updated_match,
															status=both_missing)
								match_attendance_record = Match_Attendance_Record.objects.get_or_create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=updated_match)
								session_reading_level, created = Session_Reading_Level.objects.get_or_create(
															session=session,
															user=updated_match.student)
								match.student.student_progress.level_progress.add(session_reading_level)
							# else:
							# 	print("NO ADD")
								
						else:
							updated_match.sessions_scheduled.add(session)
							match_status = Match_Status.objects.get_or_create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=updated_match,
															status=both_missing)
							match_attendance_record = Match_Attendance_Record.objects.get_or_create(
															session=session,
															match_type=scheduled_match_type,
															sch_match=updated_match)
							session_reading_level, created = Session_Reading_Level.objects.get_or_create(
															session=session,
															user=updated_match.student)
							match.student.student_progress.level_progress.add(session_reading_level)


					messages.success(request, 'Match Schedule Updated Successfully.')
					return redirect("site_admin:scheduled_matches", semester_id=semester.id,
						by_call= "Semester", by_type="Active")
				else:
					print("invalid")
					print(form.errors)
			else:	
				# print("Not POST")			
				
				form= Edit_Scheduled_Match_Form(
						initial = {
							"semester": match.semester,
							"student": match.student,
							"volunteer": match.volunteer,
							"scheduled_slots": match.scheduled_slots.all(),	
							"match_active": match.match_active,											
						}
					)
	
			context['form'] = form
			
			return render(request, "site_admin/matches/edit_match.html", context)
		else:
			return redirect('access_denied')

def post_inactive_form(request, semester):
	match_id = request.POST.get('match')
	match = Scheduled_Match.objects.get(id=match_id)
	# print("Match", match )

	status = request.POST.get('active_match')
	match.match_active = status
	match.save()

	student_user_session_status = User_Session_Status.objects.get(user=match.student) 
	student_user_session_status.current_active_match_type = None
	student_user_session_status.scheduled_match = None
	student_user_session_status.scheduled_buddy = None
	student_user_session_status.save()

	volunteer_user_session_status = User_Session_Status.objects.get(user=match.volunteer)
	volunteer_user_session_status.current_active_match_type = None
	volunteer_user_session_status.scheduled_match = None
	volunteer_user_session_status.scheduled_buddy = None
	volunteer_user_session_status.save()

	for slot in match.scheduled_slots.all():
		match.student.student_profile.scheduled_day_time_slots.remove(slot)
		match.volunteer.volunteer_profile.scheduled_day_time_slots.remove(slot)

	if status =="True":
		profile_needs_match = False
	else:
		profile_needs_match = True

	student_profile = Student_Profile.objects.get(user = match.student)
	student_profile.match_needed = profile_needs_match
	student_profile.save()

	volunteer_profile = Volunteer_Profile.objects.get(user = match.volunteer)
	volunteer_profile.match_needed = profile_needs_match
	volunteer_profile.save()

	start_date = timezone.now().date()
	end_date = semester.end_date

	future_attendance_records = Match_Attendance_Record.objects.filter(sch_match=match,
										session__date__range=(start_date, end_date))

	future_match_status = Match_Status.objects.filter(sch_match=match,
										session__date__range=(start_date, end_date))

	future_session_reading_level = Session_Reading_Level.objects.filter(
										user=match.student,
										session__date__range=(start_date, end_date))
	

	for record in future_attendance_records:
		if record.session.date == start_date:
			if record.session.session_start_date_time > timezone.now():
				record.delete()
		else:
			record.delete()

	for record in future_match_status:
		if record.session.date == start_date:
			if record.session.session_start_date_time > timezone.now():
				record.delete()
		else:
			record.delete()


	for record in future_session_reading_level:
		if record.session.date == start_date:
			if record.session.session_start_date_time > timezone.now():
				record.delete()
		else:
			record.delete()

	note_category= Note_Category.objects.get(name = "Scheduled Match Archived")
	new_note = Note.objects.create(category=note_category, author=request.user,
								   content="Match Made Inactive")
	match.notes.add(new_note)

def scheduled_matches_view(request, semester_id, by_call, by_type):
	context = {}
	context['page_title'] = "Scheduled Matches"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester
			if request.POST:
				print("Post")
				post_inactive_form(request, semester)
				return redirect("site_admin:scheduled_matches", semester_id=semester.id,
						by_call= "Semester", by_type="Active")

			else:
				if by_call == "Semester":
					context['title'] = semester.name
					context['call'] = "Semester"
					if by_type == "All":
						qs = Scheduled_Match.all_scheduled_matches_in_semester(semester)
						context['type'] = "All"
					elif by_type == "Active":
						qs = Scheduled_Match.active_scheduled_matches_in_semester(semester)
						context['type'] = "Active"
					elif by_type == "Inactive":
						qs = Scheduled_Match.inactive_scheduled_matches_in_semester(semester)
						context['type'] = "Inactive"

				elif by_call == "Session":
					if by_type == "All":
						pass
					elif by_type == "Active":
						pass
					elif by_type == "Inactive":
						pass

				context['all_matches'] = qs
				context['count'] = qs.count()

			return render(request, "site_admin/matches/view_scheduled_matches.html", context)
		else:
			return redirect('access_denied')

def sessions_by_match_view(request, match_id):
	context = {}
	context['page_title'] = "Match Scheduled Sessions"
	user = request.user
	if not user.is_authenticated:		
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)			
			match = Scheduled_Match.objects.get(id=match_id)
			match_semester = match.semester
			context['match_semester'] = match_semester
			context['match'] = match
			context['match_attendance_records'] = match.sch_match_attendance_record.all().order_by('session__date', 'session__day_time')
			context['count'] = context['match_attendance_records'].count()


			return render(request, "site_admin/matches/view_match_sessions.html", context)
		else:
			return redirect('access_denied')

def admin_home_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":

			
			context = button_bar_context_additions(context)
			return render(request, "site_admin/admin_home.html", context)
		else:
			return redirect('access_denied')

def superuser_home_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Superuser Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.is_superuser:
			
			context = button_bar_context_additions(context)	
			if request.POST:
				form = CSVBulkUploadForm(request.POST, request.FILES)
				old=Upload_CSV.objects.filter(name=request.POST['name'])
				for item in old:
					item.delete()

				if form.is_valid():
					
					obj = form.save()
					file_data = obj.csv_file.read().decode('utf-8')

					print(file_data)
					lines = file_data.split('\n')

					for line in lines:
						print("\nNew User")
						line.replace('\r', '')
						person_fields = line.split(',')
						print(person_fields)
						username=person_fields[0]
						print("Username", username)
						if username != '' and username !='username':
							if CustomUser.objects.filter(username=username).exists():
								print("Exists")
							else:
								role = Role.objects.get(name=obj.name)
								if obj.name == "Student":
									password1 = 'readbook'
									raw_password = 'readbook'
									# raw_password = "readbook"
								elif obj.name == "Volunteer":
									password1 = 'ebookbuddy'
									raw_password = 'ebookbuddy'
									# raw_password = 'ebookbuddy'
								elif obj.name == "Staff":
									password1 = 'buddystaff'
									raw_password = 'buddystaff'

								new_user = CustomUser.objects.create(
									username = username,
									first_name = person_fields[1],
									last_name = person_fields[2],
									email=person_fields[3],
									is_approved = True,
									role=role,
									password=password1)
								new_user.set_password(password1)
								new_user.save()

								create_all_user_additions(new_user)

								if new_user.role.name == "Student":
									create_student_additions(new_user, request)
									gender = Gender.objects.get(letter=person_fields[4])
									if gender !='':
										new_user.student_profile.gender = gender
									p_lang = person_fields[5]
									s_lang = person_fields[6]
									if p_lang !='':
										primary_lang = Language.objects.get(name=p_lang)
										new_user.student_profile.primary_lang = primary_lang
									if s_lang !='':
										secondary_lang = Language.objects.get(name=s_lang)
										new_user.student_profile.secondary_lang = secondary_lang

									school=person_fields[7]
									if school != '':
										user_school, created = School.objects.get_or_create(name=school)
										new_user.student_profile.school = user_school

									grade=person_fields[8]
									if grade != '':
										grade, created = Grade.objects.get_or_create(letter=grade)
										new_user.student_profile.grade = grade

									contact_person=person_fields[9]
									if contact_person != '':
										new_user.student_profile.contact_person = contact_person

									contact_number=person_fields[10]
									if contact_number != '':
										new_user.student_profile.contact_number = contact_number

									new_user.student_profile.save()

									initial_slots = person_fields[11].replace('\r', '')

									if initial_slots !='':
										initial_slots = initial_slots.replace(')', '')
										initial_slots = initial_slots.replace('(', '')
										print("Initial Slots",initial_slots)
										split_str = initial_slots.split('/')
										days= split_str[0]
										session_slot = split_str[1]
										for day in days:
											print("day", day)										
											
											day_time = Session_Day_Time.objects.get(day__letter=day, session_slot=session_slot)

											print("day_time", day_time)
											new_user.student_profile.available_day_time_slots.add(day_time)
									
								elif new_user.role.name == "Volunteer":
									create_volunteer_additions(new_user, request)
									gender = Gender.objects.get(letter=person_fields[4])
									if gender !='':
										new_user.volunteer_profile.gender = gender
									p_lang = person_fields[5]
									s_lang = person_fields[6]
									if p_lang !='':
										primary_lang = Language.objects.get(name=p_lang)
										new_user.volunteer_profile.primary_lang = primary_lang
									if s_lang !='':
										secondary_lang = Language.objects.get(name=s_lang)
										new_user.volunteer_profile.secondary_lang = secondary_lang

									contact_number=person_fields[7]
									if contact_number != '':
										new_user.volunteer_profile.contact_number = contact_number

									mega=person_fields[9]
									if mega !='':
										print("Mega", mega)
										mega_team, created = Mega_Team.objects.get_or_create(name=mega)
										new_user.volunteer_profile.mega = mega_team
									
									team=person_fields[10].replace('\r', '')
									if team !='':
										print("Team", team)
										the_leader = CustomUser.objects.get(full_name = team)
										r_team = Team.objects.get(leader=the_leader)
										new_user.volunteer_profile.team = r_team								

									new_user.volunteer_profile.save()

									initial_slots = person_fields[8]

									if initial_slots !='':
										initial_slots = initial_slots.replace(')', '')
										initial_slots = initial_slots.replace('(', '')
										print("Initial Slots",initial_slots)
										split_str = initial_slots.split('/')
										days= split_str[0]
										session_slot = split_str[1]
										for day in days:
											print("day", day)										
											
											day_time = Session_Day_Time.objects.get(day__letter=day, session_slot=session_slot)

											print("day_time", day_time)
											new_user.volunteer_profile.available_day_time_slots.add(day_time)



								elif new_user.role.name == "Staff":
									create_staff_additions(new_user, request)
					
							
							# messages.success(request, 'User Created Successfully.')
				else:
					print(form.errors)
					context['form'] = form

			else:
				form = CSVBulkUploadForm()
				context['form'] = form


			return render(request, "site_admin/superuser_home.html", context)	
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')
		else:
			return redirect('access_denied')



def create_initial_data_view(request, *args, **kwargs):
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.is_superuser:

			#Roles
			student, created = Role.objects.get_or_create(name="Student")
			volunteer, created = Role.objects.get_or_create(name="Volunteer")
			staff, created = Role.objects.get_or_create(name="Staff")

			#Relational_Engagement
			high, created = Relational_Engagement.objects.get_or_create(name="High",
				desc="High Level of Engagement and Connection; great conversation back and forth")
			medium, created = Relational_Engagement.objects.get_or_create(name="Medium",
				desc="Challenging at times but overall a good exchange of conversation")
			low, created = Relational_Engagement.objects.get_or_create(name="Low",
				desc="Great deal of silence and non-responsiveness from the student")

			#Evaluation_Level
			excellent, created = Evaluation_Level.objects.get_or_create(name="Excellent")
			good, created = Evaluation_Level.objects.get_or_create(name="Good")
			average, created = Evaluation_Level.objects.get_or_create(name="Average")
			poor, created = Evaluation_Level.objects.get_or_create(name="Poor")

			#Secondary Roles
			team_leader, created = Secondary_Role.objects.get_or_create(name="Team Leader")
			coordinator, created = Secondary_Role.objects.get_or_create(name="Coordinator")
			bilingual, created = Secondary_Role.objects.get_or_create(name="Bilingual Support")
			full_access, created = Secondary_Role.objects.get_or_create(name="Full Access")
			student_contact, created = Secondary_Role.objects.get_or_create(name="Student Contact")
			facilitator, created = Secondary_Role.objects.get_or_create(name="Facilitator")

			#Genders
			male, created = Gender.objects.get_or_create(name="Male", span="Masculino")
			female, created = Gender.objects.get_or_create(name="Female", span="Mujer")

			#Languages
			english, created = Language.objects.get_or_create(name="English")
			spanish, created = Language.objects.get_or_create(name="Spanish")

			#Days
			sunday, created = Day.objects.get_or_create(name="Sunday", span_name="Domingo",
														 number=1, short_name="Sun")
			monday, created = Day.objects.get_or_create(name="Monday", span_name="Lunes",
														 number=2, short_name="Mon")
			tuesday, created = Day.objects.get_or_create(name="Tuesday", span_name="Martes",
														 number=3, short_name="Tue")
			wednesday, created = Day.objects.get_or_create(name="Wednesday", span_name="Miércoles",
														 number=4, short_name="Wed")
			thursday, created = Day.objects.get_or_create(name="Thursday", span_name="Jueves",
													 number=5, short_name="Thu")
			friday, created = Day.objects.get_or_create(name="Friday", span_name="Viernes",
													 number=6, short_name="Fri")
			saturday, created = Day.objects.get_or_create(name="Saturday", span_name="Sábado",
													 number=7, short_name="Sat")

			#Room Type
			lobby, created = Room_Type.objects.get_or_create(name="Lobby", letter = "L")
			pending, created = Room_Type.objects.get_or_create(name="Pending", letter = "P")
			breakout, created = Room_Type.objects.get_or_create(name="Breakout", letter = "B")
			custom, created = Room_Type.objects.get_or_create(name="Custom", letter = "C")
			team_meeting, created = Room_Type.objects.get_or_create(name="Team Meeting", letter = "M")

			#Rooms
			session_lobby, created = Room.objects.get_or_create(name='Session Lobby', number = 0, room_type=lobby)
			pending_match, created = Room.objects.get_or_create(name='Match Pending', number = 0, room_type=pending)
			orientation, created = Room.objects.get_or_create(name='Orientation', number = 1, room_type=custom)

			#Breakout Rooms
			num_of_breakouts = 60

			for x in range(1, num_of_breakouts + 1):
				name = "Breakout " + str(x)
				room, created = Room.objects.get_or_create(name=name, number = x,  room_type=breakout)

			# Match Types
			scheduled, created = Match_Type.objects.get_or_create(name="Scheduled", short_name="Sch")
			temporary, created = Match_Type.objects.get_or_create(name="Temporary", short_name="Temp")

			# Temp Match Types
			sub_match, created = Temporary_Match_Type.objects.get_or_create(name="Scheduled Sub", short_name="Sub")
			in_session, created = Temporary_Match_Type.objects.get_or_create(name="In Session", short_name="I-S")


			# Note Groups
			student_profile_group, created = Note_Group.objects.get_or_create(name = "Student Profile")
			volunteer_profile_group, created = Note_Group.objects.get_or_create(name = "Volunteer Profile")
			staff_profile_group, created = Note_Group.objects.get_or_create(name = "Staff Profile")
			sch_match_group, created = Note_Group.objects.get_or_create(name = "Scheduled Match")
			temp_match_group, created = Note_Group.objects.get_or_create(name = "Temporary Match")
			student_progress_group, created = Note_Group.objects.get_or_create(name = "Student Progress")

			# Note Categories
			sch_match_created, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Scheduled Match Created")
			sch_match_edit, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Scheduled Match Edit")
			sch_match_archived, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Scheduled Match Archived")

			sch_match_br, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Both Reassigned")
			sch_match_sr, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Student Reassigned")
			sch_match_vr, created = Note_Category.objects.get_or_create(
															group=sch_match_group,
															name = "Volunteer Reassigned")



			temp_match_created, created = Note_Category.objects.get_or_create(
															group=temp_match_group,
															name = "Temporary Match Created")
			temp_match_inactive, created = Note_Category.objects.get_or_create(
															group=temp_match_group,
															name = "Temporary Match Inactive")

			
			student_profile_created, created = Note_Category.objects.get_or_create(
															group=student_profile_group,
															name = "Student Profile Created")

			student_profile_updated, created = Note_Category.objects.get_or_create(
															group=student_profile_group,
															name = "Student Profile Updated")

			volunteer_profile_created, created = Note_Category.objects.get_or_create(
															group=volunteer_profile_group,
															name = "Volunteer Profile Created")

			volunteer_profile_updated, created = Note_Category.objects.get_or_create(group=volunteer_profile_group,
																		name = "Volunteer Profile Updated")

			staff_profile_created, created = Note_Category.objects.get_or_create(
															group=staff_profile_group,
															name = "Staff Profile Created")

			staff_profile_updated, created = Note_Category.objects.get_or_create(
															group=staff_profile_group,
															name = "Staff Profile Updated")

			student_progress_created, created = Note_Category.objects.get_or_create(
													group=student_progress_group,
													name = "Student Progress Report Created")

			student_progress_updated, created = Note_Category.objects.get_or_create(
													group=student_progress_group,
													name = "Student Progress Report Updated")

			#Match Status Options
			in_room, created = Match_Status_Option.objects.get_or_create(name="In Room")
			both_r, created = Match_Status_Option.objects.get_or_create(name="Both Reassigned")
			both_m, created = Match_Status_Option.objects.get_or_create(name="Both Missing")
			pending, created = Match_Status_Option.objects.get_or_create(name="Pending Redirect")
			stu_r, created = Match_Status_Option.objects.get_or_create(name="Student Reassigned")
			vol_r, created = Match_Status_Option.objects.get_or_create(name="Volunteer Reassigned")
			stu_m, created = Match_Status_Option.objects.get_or_create(name="Student Missing")
			vol_m, created = Match_Status_Option.objects.get_or_create(name="Volunteer Missing")
			vol_m, created = Match_Status_Option.objects.get_or_create(name="Volunteer Missing")
			inac, created = Match_Status_Option.objects.get_or_create(name="Match Inactive")

			# System Messages
			next_session, created = System_Message.objects.get_or_create(
										name= "Next Session",
										eng_message="Next Session",
										span_message="Next Session Spanish")

			session_ended, created = System_Message.objects.get_or_create(
										name= "Session Ended",
										eng_message="You have missed your session for today.",
										span_message="Message in Spanish.")

			session_ns, created = System_Message.objects.get_or_create(
										name= "Sessions Not Started",
										eng_message="The Sessions for today have not started. Once the sessions begin, please click the Join Session button.",
										span_message="Spanish version")

			session_over, created = System_Message.objects.get_or_create(
										name= "Sessions Over",
										eng_message="The sessions for today have ended.",
										span_message="Spanish version")

			session_te, created = System_Message.objects.get_or_create(
										name= "Too Early",
										eng_message="Your session today has not started. Once the sessions begin, please click the Join Session button.",
										span_message="Span- Your session today has not started. Once the sessions begin, please click the Join Session button.")

			# Grades
			kinder, created = Grade.objects.get_or_create(name="Kindergarten",
														short_name="K")
			first, created = Grade.objects.get_or_create(name="1st Grade",
														short_name="1st")
			second, created = Grade.objects.get_or_create(name="2nd Grade",
														short_name="2nd")
			third, created = Grade.objects.get_or_create(name="3rd Grade",
														short_name="3rd")
			fourth, created = Grade.objects.get_or_create(name="4th Grade",
														short_name="4th")
			fifth, created = Grade.objects.get_or_create(name="5th Grade",
														short_name="5th")

			#Schools
			tuggle, created = School.objects.get_or_create(name="Tuggle Elementary School")
			ferg, created = School.objects.get_or_create(name="Ferguson Elementary School")
			brook, created = School.objects.get_or_create(name="Brookhaven Innovation Academy")
			nor, created = School.objects.get_or_create(name="Norcross Elementary School")
			unmentioned, created = School.objects.get_or_create(name="Unmentioned")
			ches, created = School.objects.get_or_create(name="Chesney Elementary School")
			fount, created = School.objects.get_or_create(name="Fountain Elementary School")
			mck, created = School.objects.get_or_create(name="McKendree Elementary School")

			#Arrival_Description
			on_time, created = Arrival_Description.objects.get_or_create(name="On Time")
			late, created = Arrival_Description.objects.get_or_create(name="Late")
			no_show, created = Arrival_Description.objects.get_or_create(name="No Show")

			#Follow_Up_Type
			tj, created = Follow_Up_Type.objects.get_or_create(name="Technical Issue")
			src, created = Follow_Up_Type.objects.get_or_create(name="Admin")
			sa, created = Follow_Up_Type.objects.get_or_create(name="Educational Specialist")
			other, created = Follow_Up_Type.objects.get_or_create(name="Misc.")

			#Mega Teams
			mega1, created = Mega_Team.objects.get_or_create(name="Alpacas")
			mega2, created = Mega_Team.objects.get_or_create(name="Owls")
			mega3, created = Mega_Team.objects.get_or_create(name="Whales")

			#Teams
			team_leaders = team_leader.roles.all()
			for item in team_leaders:
				name = item.full_name + " Team"
				team, created = Team.objects.get_or_create(name=name)
				team.leader = item
				

				room_name = name + " Meeting Room"
				new_room, created = Room.objects.get_or_create(name=room_name, number = 0, room_type=team_meeting)
				team.room = new_room
				if team.leader.role.name == "Staff":
					team.mega = team.leader.staff_profile.mega
				elif team.leader.role.name == "Volunteer":
					team.mega = team.leader.volunteer_profile.mega

				team.save()

			# number_of_teams = 15
			# for x in range(1, number_of_teams + 1):
			# 	name = "Team " + str(x)
			# 	team, created = Team.objects.get_or_create(name=name)
			# 	if x in range(1, 6):
			# 		team.mega = mega1
			# 		team.save()
			# 	elif x in range(6, 11):
			# 		team.mega = mega2
			# 		team.save()
			# 	elif x in range(11, 16):
			# 		team.mega = mega3
			# 		team.save()




			alphabets = string.ascii_uppercase

			for item in alphabets:
				letter, created = Reading_Level.objects.get_or_create(name = item)

			if user.username == "Buddy_Admin":
				staff_profile, created = Staff_Profile.objects.get_or_create(user= user)
				

				if created:
					note_category = Note_Category.objects.get(name = "Staff Profile Created")
					new_note = Note.objects.create(category=note_category, author=request.user,
													   content="Staff Profile Created")
					staff_profile.profile_notes.add(new_note)

					staff_profile.gender = female
					staff_profile.primary_lang = english
					staff_profile.save()

					note_category = Note_Category.objects.get(name = "Staff Profile Updated")
					new_note = Note.objects.create(category=note_category, author=request.user,
													   content="Staff Profile Updated")
					staff_profile.profile_notes.add(new_note)

					

				staff_session_status, created = User_Session_Status.objects.get_or_create(user= user)






			# #Note Categories
			# attendance_note, created = Note_Category.objects.get_or_create(name="Attendance")
			# student_profile, created = Note_Category.objects.get_or_create(name="Student Profile")
			# volunteer_profile, created = Note_Category.objects.get_or_create(name="Volunteer Profile")
			# student_status, created = Note_Category.objects.get_or_create(name="Student Session Status")
			# volunteer_status, created = Note_Category.objects.get_or_create(name="Volunteer Session Status")
			# scheduled_match, created = Note_Category.objects.get_or_create(name="Scheduled Match")
			# temporary_match, created = Note_Category.objects.get_or_create(name="Temporary Match")
			# archived_match, created = Note_Category.objects.get_or_create(name="Match Archived")
			# scheduled_match_created, created = Note_Category.objects.get_or_create(name="Scheduled Match Created")
			# scheduled_match_archived, created = Note_Category.objects.get_or_create(name="Scheduled Match Archived")
			# reassigned_status, created = Note_Category.objects.get_or_create(name="Reassigned Status")
			


			




			# #Session Day/Time/Slots
			# tuesdayA, created = Session_Day_Time.objects.get_or_create(day = tuesday, time_start="16:00:00",
			# 								 		time_end = "16:30:00", session_slot="A")
			# tuesdayB, created = Session_Day_Time.objects.get_or_create(day = tuesday, time_start="16:30:00",
			# 								 		time_end = "17:00:00", session_slot="B")
			# tuesdayC, created = Session_Day_Time.objects.get_or_create(day = tuesday, time_start="17:00:00",
			# 								 		time_end = "17:30:00", session_slot="C")
			# tuesdayD, created = Session_Day_Time.objects.get_or_create(day = tuesday, time_start="17:30:00",
			# 								 		time_end = "18:00:00", session_slot="D")
			# tuesdayE, created = Session_Day_Time.objects.get_or_create(day = tuesday, time_start="18:00:00",
			# 								 		time_end = "18:30:00", session_slot="E")


			# wednesdayA, created = Session_Day_Time.objects.get_or_create(day = wednesday, time_start="16:00:00",
			# 								 		time_end = "16:30:00", session_slot="A")
			# wednesdayB, created = Session_Day_Time.objects.get_or_create(day = wednesday, time_start="16:30:00",
			# 								 		time_end = "17:00:00", session_slot="B")
			# wednesdayC, created = Session_Day_Time.objects.get_or_create(day = wednesday, time_start="17:00:00",
			# 								 		time_end = "17:30:00", session_slot="C")
			# wednesdayD, created = Session_Day_Time.objects.get_or_create(day = wednesday, time_start="17:30:00",
			# 								 		time_end = "18:00:00", session_slot="D")
			# wednesdayE, created = Session_Day_Time.objects.get_or_create(day = wednesday, time_start="18:00:00",
			# 								 		time_end = "18:30:00", session_slot="E")



			# thursdayA, created = Session_Day_Time.objects.get_or_create(day = thursday, time_start="16:00:00",
			# 								 		time_end = "16:30:00", session_slot="A")
			# thursdayB, created = Session_Day_Time.objects.get_or_create(day = thursday, time_start="16:30:00",
			# 								 		time_end = "17:00:00", session_slot="B")
			# thursdayC, created = Session_Day_Time.objects.get_or_create(day = thursday, time_start="17:00:00",
			# 								 		time_end = "17:30:00", session_slot="C")
			# thursdayD, created = Session_Day_Time.objects.get_or_create(day = thursday, time_start="17:30:00",
			# 								 		time_end = "18:00:00", session_slot="D")
			# thursdayE, created = Session_Day_Time.objects.get_or_create(day = thursday, time_start="18:00:00",
			# 								 		time_end = "18:30:00", session_slot="E")

			


			# #User Locations
			# users = CustomUser.objects.all()
			# for user in users:
			# 	location, created = User_Room_Location.objects.get_or_create(user=user)
			# 	location.room=session_lobby
			# 	location.in_room = False  
			# 	location.save()


			# #Students
			# students = CustomUser.objects.filter(role__name="Student")
			# for student in students:
			# 	student_profile, created = Student_Profile.objects.get_or_create(user=student)
			# 	student_session_status, created = Student_Session_Status.objects.get_or_create(user=student)

			# #volunteers
			# volunteers = CustomUser.objects.filter(role__name="Volunteer")
			# for volunteer in volunteers:
			# 	volunteer_profile, created = Volunteer_Profile.objects.get_or_create(user=volunteer)
			# 	volunteer_session_status, created = Volunteer_Session_Status.objects.get_or_create(user=volunteer)



			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')
		else:
			return redirect('access_denied')
#Done

def create_student_additions(user, request):
	#Student_Profile
	
	profile = Student_Profile.objects.create(user=user, match_needed = True)
	note_category = Note_Category.objects.get(name = "Student Profile Created")
	new_note = Note.objects.create(category=note_category, author=request.user,
												   content="Student User & Profile Created")
	profile.profile_notes.add(new_note)

	progress = Student_Progress.objects.create(user=user)

	note_category2 = Note_Category.objects.get(name = "Student Progress Report Created")
	new_note2 = Note.objects.create(category=note_category2, author=request.user,
												   content="Student Progress Report Created")
	progress.progress_notes.add(new_note2)

	student_report = Student_Report.objects.create(user=user, semester=active_semester())

def create_volunteer_additions(user, request):
	#Volunteer_Profile
	profile = Volunteer_Profile.objects.create(user=user, match_needed = True)
	note_category = Note_Category.objects.get(name = "Volunteer Profile Created")
	new_note = Note.objects.create(category=note_category, author=request.user,
												   content="Volunteer User & Profile Created")
	profile.profile_notes.add(new_note)
	volunteer_report = Volunteer_Report.objects.create(user=user, semester=active_semester())

def create_staff_additions(user, request):
	#Staff_Profile
	profile = Staff_Profile.objects.create(user=user)
	note_category = Note_Category.objects.get(name = "Staff Profile Created")
	new_note = Note.objects.create(category=note_category, author=request.user,
												   content="Staff User & Profile Created")
	profile.profile_notes.add(new_note)

def create_all_user_additions(user):
	#User_Session_Status
	session_status = User_Session_Status.objects.create(user=user)
	create_avatar(user)

def create_user_view(request):
	context = {}
	context['page_title'] = "Create User"

	user = request.user	

	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.role.name == "Staff":
			# context['active_semester'] = active_semester()				
			if request.POST:
				form = Create_User_Form(request.POST)

				if form.is_valid():
					form.save()
					email = form.cleaned_data.get('email').lower()
					username = form.cleaned_data.get('username').lower()
					raw_password = form.cleaned_data.get('password1')
					new_user = authenticate(username=username, password=raw_password)
					
					role = Role.objects.get(id=request.POST.get('role'))
					new_user.role = role
					new_user.save()

					create_all_user_additions(new_user)

					if new_user.role.name == "Student":
						create_student_additions(new_user, request)
					elif new_user.role.name == "Volunteer":
						create_volunteer_additions(new_user, request)
					elif new_user.role.name == "Staff":
						create_staff_additions(new_user, request)
					messages.success(request, 'User Created Successfully.')
				else:
					print(form.errors)
					context['registration_form'] = form

			else:
				form = Create_User_Form()
				context['registration_form'] = form
			context = button_bar_context_additions(context)
			return render(request, 'site_admin/users/create_user.html', context)
		else:
			return redirect('access_denied')

def edit_user_view(request, user_id):
	context = {}
	context['page_title'] = "Edit User"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		edit_user = CustomUser.objects.get(id=user_id)
		if user.role.name == "Staff":
			member = CustomUser.objects.get(id=user_id)
			if request.POST:
				form = User_Update_Form(request.POST, instance=member)
				
				if form.is_valid():
					# print("valid")
					obj = form.save()
					create_avatar(obj)
					messages.success(request, 'User Updated Successfully.')
					return redirect('site_admin:users', role_name="All")
				else:
					# print("invalid")
					print(form.errors)
			else:				
				
				form= User_Update_Form(
						initial = {
								"first_name": member.first_name,
								"middle_name": member.middle_name,
								"last_name": member.last_name,
								"username": member.username,
								"email": member.email,													
						}
					)
	
			context['member'] = member
			context['edit_user'] = edit_user
			context['form'] = form

			context = button_bar_context_additions(context)
			return render(request, 'site_admin/users/edit_user.html', context)
		else:
			return redirect('access_denied')

def current_sessions_by_day_semester(semester):
	days_with_sessions = Day_With_Daily_Session.objects.filter(semester = semester)
	return days_with_sessions

def days_with_sessions_by_semester_view(request, semester_id, order_by):
	context = {}
	context['page_title'] = "Sessions By Day"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester
			if order_by == "All":
				context['days_in_semester'] = current_sessions_by_day_semester(semester).order_by('date')
			else:
				context['days_in_semester'] = current_sessions_by_day_semester(semester).order_by(order_by, 'date')

			context['count'] = current_sessions_by_day_semester(semester).count()
			return render(request, "site_admin/sessions/sessions_by_day.html", context)
		else:
			return redirect('access_denied')

def edit_sessions_for_semester_view(request, semester_id):
	context = {}
	context['page_title'] = "Edit Sessions"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context['active_semester'] = active_semester()
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester
			
			day_times = Session_Day_Time.objects.all()
			context['day_times'] = day_times

			if request.POST:
				semester_id = request.POST.get('semester')
				semester = Semester.objects.get(id=semester_id)
				day_time_slots = request.POST.getlist('day_time')
				# print("day_time_slots", day_time_slots)
				create_semester_sessions(semester, day_time_slots)
				messages.success(request, 'Sessions Updated Successfully.')
				return redirect('site_admin:semesters')
			# else:
			# 	print("Not Post")

			return render(request, "site_admin/sessions/edit_sessions_for_semester.html", context)
		else:
			return redirect('access_denied')

def current_sessions_by_semester(semester):
	sessions_by_semester = Daily_Session.objects.filter(semester = semester,
														archive_session=False)
	return sessions_by_semester

def create_semester_sessions(semester, slots):
	start_date =semester.start_date
	end_date = semester.end_date
	delta = datetime.timedelta(days=1)
	semester.day_time_slots.clear()
	active_daily_sessions = []

	week = 1
	
	while start_date <= end_date:		
		start_day = start_date.weekday()

		if start_day == 6:
			week = week + 1
		# else:
		# 	print(start_day)

		start_day_of_week = calendar.day_name[start_day]
		day = start_date

		for entry in slots:
			slot = Session_Day_Time.objects.get(id=int(entry))
			semester.day_time_slots.add(slot)
			semester.days.add(slot.day)

			day_name  = slot.day.name
			if start_day_of_week == day_name:
				# print("Day", day, start_day_of_week)			# 
				session = Daily_Session.objects.get_or_create(semester = semester, day_time = slot, date=day, week = week)
				active_daily_sessions.append(session[0])
		start_date += delta

	current_semester_sessions = Daily_Session.objects.filter(semester = semester)

	for session in current_semester_sessions:
		if session in active_daily_sessions:
			session.archive_session = False
		else:
			session.archive_session = True
		session.save()	

def sessions_by_semester_view(request, semester_id, order_by):
	context = {}
	context['page_title'] = "Sessions"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester
			if order_by == "All":
				context['sessions_in_semester'] = current_sessions_by_semester(semester).order_by('date', 'day_time__session_slot')
			else:
				context['sessions_in_semester'] = current_sessions_by_semester(semester).order_by(order_by, 'date', 'day_time__session_slot')

			context['count'] = current_sessions_by_semester(semester).count()
			return render(request, "site_admin/sessions/sessions_by_semester.html", context)
		else:
			return redirect('access_denied')


def create_sessions_for_semester_view(request, semester_id):
	context = {}
	context['page_title'] = "Create Sessions"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester
			
			day_times = Session_Day_Time.objects.all()
			context['day_times'] = day_times

			if request.POST:
				semester_id = request.POST.get('semester')
				day_time_slots = request.POST.getlist('day_time')
				# print("day_time_slots", day_time_slots)
				create_semester_sessions(semester, day_time_slots)
				messages.success(request, 'Sessions Created Successfully.')
				return redirect('site_admin:semesters')
			# else:
			# 	print("Not Post")

			return render(request, "site_admin/sessions/create_sessions_for_semester.html", context)
		else:
			return redirect('access_denied')

def edit_semester_view(request, semester_id):
	context = {}
	context['page_title'] = "Edit Semester"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)				
			
			semester = Semester.objects.get(id=semester_id)
			context['semester'] = semester

			if request.POST:
				form = Edit_Semester_Form(request.POST, instance=semester)
				day_time_slots = request.POST.getlist('day_time_slots')
				days = request.POST.getlist('days')

				if form.is_valid():
					obj = form.save()

					day_time_slots_current = []
					for slot in obj.day_time_slots.all():
						day_time_slots_current.append(slot.id)

					
					create_semester_sessions(obj, day_time_slots_current)

					messages.success(request, 'Semester Updated Successfully.')
					return redirect('site_admin:semesters')
				else:
					print(form.errors)
					

			else:				
				form= Edit_Semester_Form(
						initial = {
								"name": semester.name,
								"start_date": str(semester.start_date),
								"end_date": str(semester.end_date),
								"active_semester": semester.active_semester,

						}
					)
			
			context['form'] = form
			
			return render(request, "site_admin/semesters/edit_semester.html", context)
		else:
			return redirect('access_denied')


def create_custom_room_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Create Room"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			custom_room_types = Room_Type.objects.filter(Q(letter="C") | Q(letter="M"))
			context['room_types'] = custom_room_types
			# Room_Type.objects.filter(letter = "C")
			
			# context['room_type'] = custom_room_type
			if request.POST:
				form = Create_Room_Form(request.POST)
				if form.is_valid():
					form.save()
					messages.success(request, 'Room Created Successfully.')
					return redirect('site_admin:all_rooms')
				else:
					print(form.errors)
					context['form'] = form

			else:
				form = Create_Room_Form()
				context['form'] = form
			
			return render(request, "site_admin/rooms/create_custom_room.html", context)
		else:
			return redirect('access_denied')

def create_additional_breakout_rooms_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Create Additional Breakout Rooms"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)				
			if request.POST:
				form = Breakout_Rooms_Form(request.POST)
				breakout_room_type = Room_Type.objects.get(letter = "B")
				
				if form.is_valid():
					last_breakout_room = Room.objects.filter(room_type=breakout_room_type).last()
					starting_number = last_breakout_room.number + 1
					num_of_breakouts = int(request.POST.get('number_of_rooms'))
					for x in range(1, num_of_breakouts + 1):
						name = "Breakout " + str(starting_number)
						Room.objects.get_or_create(name=name, number = starting_number,  room_type=breakout_room_type)

						starting_number = starting_number + 1
					messages.success(request, 'Breakout Rooms Created Successfully.')
					return redirect('site_admin:all_rooms')
				else:
					print(form.errors)
					context['form'] = form

			else:
				# print("else")
				form = Breakout_Rooms_Form()
				context['form'] = form
			
			return render(request, "site_admin/rooms/create_breakout_rooms.html", context)
		else:
			return redirect('access_denied')

def all_rooms():
	all_rooms = Room.objects.all()
	return all_rooms

def all_rooms_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Rooms"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			context['all_rooms'] = all_rooms()
			context['count'] = all_rooms().count()
			return render(request, "site_admin/rooms/all_rooms.html", context)
		else:
			return redirect('access_denied')

def get_users_queryset(query=None):
	queryset = []
	queries = query.split(" ")

	for q in queries:
		users = CustomUser.objects.filter(
			Q(first_name__icontains=q)|
			Q(middle_name__icontains=q)|
			Q(last_name__icontains=q)|
			Q(email__icontains=q)|
			Q(username__icontains=q)).order_by('username')

		for user in users:
			queryset.append(user)

	return list(set(queryset))

def all_users():
	all_users = CustomUser.objects.all().order_by('username')
	return all_users

def all_students():
	role = Role.objects.get(name="Student")
	all_students = CustomUser.objects.filter(role = role.id).order_by('username')
	return all_students

def all_volunteers():
	role = Role.objects.get(name="Volunteer")
	all_volunteers = CustomUser.objects.filter(role= role.id).order_by('username')
	return all_volunteers

def all_staff():
	role = Role.objects.get(name="Staff")
	all_staff = CustomUser.objects.filter(role = role.id).order_by('username')
	return all_staff

def users_by_role_view(request, role_name, **kwargs):
	context = {}
	context['page_title'] = "Users"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			if request.GET:
				query = request.GET.get('q', '')
				type_user = request.GET.get('type_user', '')
				if query !='':
					context['users'] = get_users_queryset(query)
					context['count'] = len(context['users'])
					context['role_name'] = "Matching"
				else:
					context['users'] = all_users()
					context['count'] = context['users'].count
					context['role_name'] = "All"

						
			else:
				if role_name == "All":
					context['users'] = CustomUser.objects.all().order_by('username')
					context['count'] = context['users'].count
					context['role_name'] = role_name
				else:
					context['users'] = CustomUser.objects.filter(role__name=role_name).order_by('username')
					context['count'] = context['users'].count
					context['role_name'] = role_name

			
			context = button_bar_context_additions(context)	
			return render(request, "site_admin/users/view_users.html", context)
		else:
			return redirect('access_denied')

def all_semesters():
	all_semesters = Semester.objects.all()
	return all_semesters

def active_semester():
	if Semester.objects.filter(active_semester=True).exists():
		active_semester = Semester.objects.get(active_semester=True)
		return active_semester
	else:
		return None

def profile_context_additions(context):
	slots = all_day_time_slots()
	context['slots'] = slots

	genders = Gender.objects.all()
	context['genders'] = genders

	languages = Language.objects.all()
	context['languages'] = languages

	schools = School.objects.all().order_by('name')
	context['schools'] = schools

	grades = Grade.objects.all().order_by('id')
	context['grades'] = grades

	teams = Team.objects.all().order_by('id')
	context['teams'] = teams

	mega_teams = Mega_Team.objects.all().order_by('id')
	context['mega_teams'] = mega_teams

	secondary_roles = Secondary_Role.objects.all().order_by('name')
	context['secondary_roles'] = secondary_roles

	return context


def edit_staff_profile_view(request, user_id):
	context = {}
	context['page_title'] = "Edit Staff Profile"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else: 		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)			
			context = profile_context_additions(context)			
			member = CustomUser.objects.get(id=user_id)
			context['member']  = member
			staff_profile = Staff_Profile.objects.get(user=member)
			current_notes = staff_profile.profile_notes.all()

			if request.POST:
				form = Edit_Staff_Profile_Form(request.POST, instance=staff_profile)
				
				if form.is_valid():					
					obj = form.save()

					notes = request.POST.getlist('profile_note')
					additional_roles = request.POST.getlist('role2')

					for item in notes:
						obj.profile_notes.add(item)

					note_category = Note_Category.objects.get(name = "Staff Profile Updated")
					new_note = Note.objects.create(category=note_category, author=request.user,
																   content="Staff Profile Updated")
					obj.profile_notes.add(new_note)

					team_leader = Secondary_Role.objects.get(name = "Team Leader")
					coordinator = Secondary_Role.objects.get(name = "Coordinator")
					# print(team_leader.id, coordinator.id)

					member.secondary_roles.clear()
					for role2 in additional_roles:
						member.secondary_roles.add(role2)
						print(role2, coordinator)
						if int(role2) == coordinator.id:
							obj.mega.coordinator = member
							obj.mega.save()
							
						elif int(role2) == team_leader.id:
							obj.team.leader = member
							obj.team.save()

					if coordinator not in member.secondary_roles.all():
						mega_teams = Mega_Team.objects.filter(coordinator=member)
						for team in mega_teams:
							team.coordinator = None
							team.save()
							

					if team_leader not in member.secondary_roles.all():
						member_teams = Team.objects.filter(leader=member)
						for team in member_teams:
							team.leader = None
							team.save()

					success_message = member.full_name + ' - Profile Updated Successfully.'
					messages.success(request, success_message)
					# return redirect('site_admin:user_profile', user_id=member.id)
					return redirect('site_admin:users', role_name="All")
				else:
					# print("invalid")
					print(form.errors)
					context['errors'] = form.errors.as_ul

				
			form= Edit_Staff_Profile_Form(
					initial = {
							"user": staff_profile.user.id,
							"gender": staff_profile.gender,
							"primary_lang": staff_profile.primary_lang,
							"secondary_lang": staff_profile.secondary_lang,
							"contact_number": staff_profile.contact_number,
							"comment": staff_profile.comment,
							"mega": staff_profile.mega,
							"team": staff_profile.team,
							"date_created": staff_profile.date_created,
							"last_updated": staff_profile.last_updated,													
					}
				)
	
			context['member'] = member
			context['form'] = form
			context['current_notes'] = current_notes
			
			return render(request, 'site_admin/users/edit_staff_profile.html', context)
		else:
			return redirect('access_denied')

def edit_student_profile_view(request, user_id):
	context = {}
	context['page_title'] = "Edit Student Profile"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else: 
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)			
			context = profile_context_additions(context)			
			member = CustomUser.objects.get(id=user_id)
			context['member']  = member
			student_profile = Student_Profile.objects.get(user=member)
			current_notes = student_profile.profile_notes.all()


			if request.POST:
				form = Edit_Student_Profile_Form(request.POST, instance=student_profile)				
				if form.is_valid():					
					obj = form.save()

					notes = request.POST.getlist('profile_note')


					for item in notes:
						obj.profile_notes.add(item)

					note_category = Note_Category.objects.get(name = "Student Profile Updated")
					new_note = Note.objects.create(category=note_category, author=request.user,
																   content="Student Profile Updated")
					obj.profile_notes.add(new_note)

					available_slots = request.POST.getlist('day_time')
					sch_slots = request.POST.getlist('sch_day_time')

					# print("available_slots", available_slots)					
					obj.available_day_time_slots.clear()
					for slot in available_slots:
						obj.available_day_time_slots.add(slot)

					obj.scheduled_day_time_slots.clear()
					for slot in sch_slots:
						obj.scheduled_day_time_slots.add(slot)

					success_message = member.full_name + ' - Profile Updated Successfully.'
					messages.success(request, success_message)
					# return redirect('site_admin:user_profile', user_id=member.id)
					return redirect('site_admin:users', role_name="All")
				else:
					# print("invalid")
					print(form.errors)
					context['errors'] = form.errors.as_ul

				
			form= Edit_Student_Profile_Form(
					initial = {
							"user": student_profile.user.id,
							"match_needed": student_profile.match_needed,
							"gender": student_profile.gender,
							"primary_lang": student_profile.primary_lang,
							"secondary_lang": student_profile.secondary_lang,
							"school": student_profile.school,
							"grade": student_profile.grade,
							"contact_person": student_profile.contact_person,
							"contact_relationship": student_profile.contact_relationship,
							"contact_number": student_profile.contact_number,
							"comment": student_profile.comment,
							"date_created": student_profile.date_created,
							"last_updated": student_profile.last_updated,													
					}
				)
	
			context['member'] = member
			context['form'] = form
			context['current_notes'] = current_notes

			return render(request, 'site_admin/users/edit_student_profile.html', context)
		else:
			return redirect('access_denied')


def edit_volunteer_profile_view(request, user_id):
	context = {}
	context['page_title'] = "Edit Volunteer Profile"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)			
			context = profile_context_additions(context)			
			member = CustomUser.objects.get(id=user_id)
			context['member']  = member
			volunteer_profile = Volunteer_Profile.objects.get(user=member)
			current_notes = volunteer_profile.profile_notes.all()

			if request.POST:
				form = Edit_Volunteer_Profile_Form(request.POST, instance=volunteer_profile)				
				if form.is_valid():					
					obj = form.save()

					notes = request.POST.getlist('profile_note')

					for item in notes:
						obj.profile_notes.add(item)

					note_category = Note_Category.objects.get(name = "Volunteer Profile Updated")
					new_note = Note.objects.create(category=note_category, author=request.user,
																   content="Volunteer Profile Updated")
					obj.profile_notes.add(new_note)

					available_slots = request.POST.getlist('day_time')
					sch_slots = request.POST.getlist('sch_day_time')
					additional_roles = request.POST.getlist('role2')

					obj.available_day_time_slots.clear()
					for slot in available_slots:
						obj.available_day_time_slots.add(slot)

					obj.scheduled_day_time_slots.clear()
					for slot in sch_slots:
						obj.scheduled_day_time_slots.add(slot)

					team_leader = Secondary_Role.objects.get(name = "Team Leader")
					coordinator = Secondary_Role.objects.get(name = "Coordinator")
					# print(team_leader.id, coordinator.id)

					member.secondary_roles.clear()
					for role2 in additional_roles:
						member.secondary_roles.add(role2)
						print(role2, coordinator)
						if int(role2) == coordinator.id:
							obj.mega.coordinator = member
							obj.mega.save()
							
						elif int(role2) == team_leader.id:
							obj.team.leader = member
							obj.team.save()

					if coordinator not in member.secondary_roles.all():
						mega_teams = Mega_Team.objects.filter(coordinator=member)
						for team in mega_teams:
							team.coordinator = None
							team.save()
							

					if team_leader not in member.secondary_roles.all():
						member_teams = Team.objects.filter(leader=member)
						for team in member_teams:
							team.leader = None
							team.save()





					success_message = member.full_name + ' - Profile Updated Successfully.'
					messages.success(request, success_message)
					# return redirect('site_admin:user_profile', user_id=member.id)
					return redirect('site_admin:users', role_name="All")
				else:
					# print("invalid")
					print(form.errors)
					context['errors'] = form.errors.as_ul

				
			form= Edit_Volunteer_Profile_Form(
					initial = {
							"user": volunteer_profile.user.id,
							"match_needed": volunteer_profile.match_needed,
							"mega": volunteer_profile.mega,
							"team": volunteer_profile.team,
							"gender": volunteer_profile.gender,
							"primary_lang": volunteer_profile.primary_lang,
							"secondary_lang": volunteer_profile.secondary_lang,
							"contact_number": volunteer_profile.contact_number,
							"comment": volunteer_profile.comment,
							"date_created": volunteer_profile.date_created,
							"last_updated": volunteer_profile.last_updated,													
					}
				)
	
			context['member'] = member
			context['form'] = form
			context['current_notes'] = current_notes

			return render(request, 'site_admin/users/edit_volunteer_profile.html', context)
		else:
			return redirect('access_denied')

def make_semester_active(request, semester_id):
	# print("here")
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			# context['active_semester'] = active_semester()
			semester = get_object_or_404(Semester, id=semester_id)
			semesters = Semester.objects.all()
			for sem in semesters:
				sem.active_semester = False
				sem.save()
			semester.active_semester = True
			semester.save()
			return redirect('site_admin:semesters')
		else:
			return redirect('access_denied')

def view_profile_view(request, user_id):
	context = {}	
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		
		if user.role.name == "Staff":			
			member = CustomUser.objects.get(id=user_id)
			context['member'] = member
			context = button_bar_context_additions(context)

			if member.role.name == "Staff":
				context['page_title'] = "Staff Profile"
				return render(request, 'site_admin/users/view_staff_profile.html', context)
			elif member.role.name == "Student":
				context['page_title'] = "Student Profile"
				sch_matches = Scheduled_Match.objects.filter(student=member).order_by('-match_active')
				context['scheduled_matches'] = sch_matches
				all_levels = Reading_Level.objects.all().order_by('id')
				context['all_levels'] = all_levels
				return render(request, 'site_admin/users/view_student_profile.html', context)
			elif member.role.name == "Volunteer":
				context['page_title'] = "Volunteer Profile"
				sch_matches = Scheduled_Match.objects.filter(volunteer=member).order_by('-match_active')
				context['scheduled_matches'] = sch_matches
				return render(request, 'site_admin/users/view_volunteer_profile.html', context)
			
		else:
			return redirect('access_denied')

def all_semesters_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Semesters"
	context = button_bar_context_additions(context)
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context['all_semesters'] = all_semesters()
			context['count'] = all_semesters().count()		
			return render(request, "site_admin/semesters/view_semesters.html", context)
		else:
			return redirect('access_denied')

def create_semester_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Create Semester"
	context = button_bar_context_additions(context)
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			if request.POST:
				form = Create_Semester_Form(request.POST)

				if form.is_valid():
					form.save()
					messages.success(request, 'Semester Created Successfully.')
					return redirect('site_admin:semesters')
				else:
					print(form.errors)
					context['form'] = form

			else:
				form = Create_Semester_Form()
				context['form'] = form
			
			return render(request, "site_admin/semesters/create_semester.html", context)
		else:
			return redirect('access_denied')

def all_day_time_slots():
	all_dts = Session_Day_Time.objects.all()
	return all_dts

def all_day_time_slots_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Day|Time|Slots"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			context['all_dts'] = all_day_time_slots()
			context['count'] = all_day_time_slots().count()
			return render(request, "site_admin/sessions/all_session_day_time_slots.html", context)
		else:
			return redirect('access_denied')

def create_day_time_slot_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Create Day/Time/Slot"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)
			days = Day.objects.all()
			context['days'] = days				

			if request.POST:
				form = Create_Session_Day_Time_Form(request.POST)

				if form.is_valid():
					form.save()
					messages.success(request, 'Day/Time/Slot Created Successfully.')
					return redirect('site_admin:day_time_slots')
				else:
					print("Not Valid")
					print(form.errors)
					context['form'] = form

			else:
				form = Create_Session_Day_Time_Form()
				context['form'] = form
			
			return render(request, "site_admin/sessions/add_session_day_time_slot.html", context)
		else:
			return redirect('access_denied')

def edit_day_time_slot_view(request, entry_id):
	context = {}
	context['page_title'] = "Edit Day/Time/Slot"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			entry = Session_Day_Time.objects.get(id=entry_id)
			days = Day.objects.all()
			context['days'] = days	

			if request.POST:
				form = Create_Session_Day_Time_Form(request.POST, instance=entry)
				
				if form.is_valid():
					obj = form.save()
					sessions = Daily_Session.objects.filter(day_time=obj)
					for session in sessions:
						session.save()

					messages.success(request, 'Entry Updated Successfully.')
					return redirect('site_admin:day_time_slots')
				else:
					print(form.errors)
					context['errors'] = form.errors.as_ul
			else:
				form= Create_Session_Day_Time_Form(
						initial = {
								"day": entry.day.id,
								"time_start": str(entry.time_start),
								"time_end": str(entry.time_end),
								"session_slot": entry.session_slot,
								"currently_active": entry.currently_active,													
						}
					)
	
			context['entry'] = entry
			context['form'] = form			
			return render(request, 'site_admin/sessions/edit_session_day_time_slot.html', context)
		else:
			return redirect('access_denied')

def make_slot_active(request, entry_id):
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			# context['active_semester'] = active_semester()
			slot = get_object_or_404(Session_Day_Time, id=entry_id)
			if slot.currently_active == True:
				slot.currently_active = False;
			elif slot.currently_active == False:
				slot.currently_active = True;
			slot.save()

			return redirect('site_admin:day_time_slots')
		else:
			return redirect('access_denied')

#AJAX CALLS INTERNAL
def approve_user(request):
	response = {}
	# request should be ajax and method should be GET.
	if request.is_ajax and request.method == "GET":
		# get from the client side.		
		user_id = request.GET.get("target_id", None)

		# check database.
		if CustomUser.objects.filter(id = user_id).exists():
			response['valid'] = True
			user = get_object_or_404(CustomUser, id=user_id)

			if user.is_approved == True:
				user.is_approved = False;
			elif user.is_approved == False:
				user.is_approved = True;
			user.save()

			# print("last_updated", user.last_updated)
			last_updated = user.last_updated.strftime("%Y-%m-%d %H:%M:%S")
			
			response['status'] = user.is_approved
			response['last_updated'] = last_updated
			# return redirect('site_admin:users')
			return HttpResponse(dumps(response), content_type="application/json")
		else:
			response['valid'] = False
			return HttpResponse(dumps(response), content_type="application/json")

	return JsonResponse({}, status = 400)



def create_team_meetings_view(request):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.is_superuser:
			print("\n\n\nCreating Team Meetings")
			teams = Team.objects.all()
			semester = active_semester()

			for team in teams:
				team_meeting = Team_Meeting.objects.get_or_create(team=team,
																day=team.meeting_day, 
																time=team.meeting_time)

			team_meetings = Team_Meeting.objects.all()

			days = ['Monday', 'Thursday']

			start_date =semester.start_date
			end_date = semester.end_date
			delta = datetime.timedelta(days=1)
			
			
	
			while start_date <= end_date:		
				start_day = start_date.weekday()


				start_day_of_week = calendar.day_name[start_day]
				day = start_date
				if start_day_of_week in days:
					print("Date", start_date, start_day_of_week)
					the_day = Day.objects.get(name=start_day_of_week)
					day_with_meeting, created = Day_With_Team_Meeting.objects.get_or_create(date=start_date, day=the_day, semester = semester)
					for meeting in team_meetings:
						if meeting.day == the_day:
							day_with_meeting.day_meetings.add(meeting)

				start_date += delta
						
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')	

		else:
			return redirect('access_denied')

def all_days_attendance_reports_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Attendance Reports"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			all_days = Day_With_Daily_Session.objects.all()
			context['all_days'] = all_days

			return render(request, "site_admin/reports/all_days_attendance.html", context)
		else:
			return redirect('access_denied')


def all_session_evaluations_reports_view(request, *args, **kwargs):
	context = {}
	context['page_title'] = "Session Evaluations"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.role.name == "Staff":
			context = button_bar_context_additions(context)	
			
			all_evals = End_Session_Evaluation.objects.all().order_by('date')
			context['all_evals'] = all_evals

			return render(request, "site_admin/reports/session_evaluations_report.html", context)
		else:
			return redirect('access_denied')


def get_ajax_queryset(role, query=None):
	print("get_ajax_queryset", role, query)

	queryset = []
	queries = query.split(" ")
	print("queries", queries)

	for q in queries:
		users = CustomUser.objects.filter(Q(role__name__icontains=role),
			Q(first_name__icontains=q)|
			Q(middle_name__icontains=q)|
			Q(last_name__icontains=q)|
			Q(email__icontains=q)|
			Q(username__icontains=q)).exclude(username__contains=role)

		for user in users:
			queryset.append(user)

	return list(set(queryset))

def ajax_search_users(request):
	response = {}
	# request should be ajax and method should be GET.
	if request.is_ajax and request.method == "GET":
		# get from the client side.		
		q = request.GET.get("search_terms", None)
		role = request.GET.get("role", None)

		

		# check database.
		if Role.objects.filter(name = role).exists():
			total_users = CustomUser.objects.filter(role__name=role).exclude(username__contains=role)
			response['valid'] = True

			results = get_ajax_queryset(role, q)

			results_id_list = []

			for result in results:
				results_id_list.append(result.id)

			response['total_users'] = total_users.count()
			response['results_id_list'] = results_id_list
			response['count'] = len(results_id_list)

			return HttpResponse(dumps(response), content_type="application/json")
		else:
			response['valid'] = False
			return HttpResponse(dumps(response), content_type="application/json")

	return JsonResponse({}, status = 400)



def create_avatars_view(request):
	context = {}
	context['page_title'] = "Admin Home"
	user = request.user
	if not user.is_authenticated:
		return redirect('must_authenticate')
	else:		
		if user.is_superuser:
			all_user_avatars = CustomUser.objects.all()
			for item in all_user_avatars:
				create_avatar(item)
			
						
			return redirect('site_admin:superuser_home')
		elif user.role.name == "Staff":	
			return redirect('site_admin:admin_home')	

		else:
			return redirect('access_denied')